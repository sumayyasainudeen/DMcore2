from django.shortcuts import render
from venv import create
import qrcode
import random
import os, json, math
# import psycopg2
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from urllib.parse import urlencode
from django.views.decorators.csrf import csrf_exempt
from django. contrib import messages
from unicodedata import name
# pip install openpyxl
from openpyxl import Workbook

from django.shortcuts import render, redirect
from .models import *
from datetime import datetime,date, timedelta
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models import Q
from io import BytesIO
from django.core.files import File
from django.conf import settings
from django.db.models import Q
from num2words import num2words
from django.http import JsonResponse
from django.core.mail import send_mail

from django.core.files.storage import FileSystemStorage

from django.http import HttpResponse
from django.template.loader import get_template
# from xhtml2pdf import pisa


from django.views.decorators.http import require_GET

import requests

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect
# Create your views here.

from datetime import datetime, time, timedelta
from django.db.models.functions import TruncHour
from datetime import time

from datetime import datetime, timedelta, time
from django.shortcuts import render
from openpyxl import load_workbook

from schedule import every
from django.views.generic import View
from django.utils import timezone

# import analytics
#----------------------------------------------------------Login, Sign Up, Reset, Internshipform 
def login(request):
    return render(request, 'home/login.html')

def signin(request):
    
    if request.method == 'POST':
        email  = request.POST['email']
        password = request.POST['password']
        user = authenticate(username=email, password=password)
        if user is not None:
            return redirect('login')
        
        if user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Admin",status="active").exists():

            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            
            request.session['userid'] = member.id
            
            return redirect('ad_profile')


        elif user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Digital Marketing Head",status="active").exists():
            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            request.session['userid'] = member.id
            return redirect('he_profile')

        elif user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Team Lead",status="active").exists():
            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            request.session['userid'] = member.id
            return redirect('tl_profile')

        elif user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Digital Marketing Executive",status="active").exists():
            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            request.session['userid'] = member.id

            return redirect('ex_profile')
        
        elif user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Data Manager",status="active").exists():
            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            request.session['userid'] = member.id
            return redirect('dm_profile')
        
        elif user_registration.objects.filter(email=request.POST['email'], password=request.POST['password'],department="Telecaller",status="active").exists():
            member = user_registration.objects.get(email=request.POST['email'],password=request.POST['password'])
            request.session['userid'] = member.id

            return redirect('tc_profile')
        
        else:
           
            return redirect('login')


           

def signup(request):
    return render(request, 'home/signup.html')

def registration_form(request):

 
    a = user_registration()
    b = qualification()
    c = extracurricular()

    if request.method == 'POST':
        if  user_registration.objects.filter(email=request.POST['email']).exists():
            
            msg_error = "Mail id already exist"
            return render(request, 'home/signup.html',{'msg_error': msg_error})
        else:
            
            a.fullname = request.POST['fname']
            a.fathername = request.POST['fathername']
            a.mothername = request.POST['mothername']
            a.dateofbirth = request.POST['dob']
            a.gender = request.POST['gender']
            a.presentaddress1 = request.POST['address1']
            a.presentaddress2  =  request.POST['address2']
            a.presentaddress3 =  request.POST['address3']
            a.pincode = request.POST['pincode']
            a.district  =  request.POST['district']
            a.state  =  request.POST['state']
            a.country  =  request.POST['country']
            a.permanentaddress1 = request.POST['paddress1']
            a.permanentaddress2  =  request.POST['paddress2']
            a.permanentaddress3  =  request.POST['paddress3']
            a.permanentpincode = request.POST['ppincode']
            a.permanentdistrict  =  request.POST['pdistrict']
            a.permanentstate  =  request.POST['pstate']
            a.permanentcountry =  request.POST['pcountry']
            a.mobile = request.POST['mobile']
            a.alternativeno = request.POST['alternative']
            a.department = request.POST['department']
            a.email = request.POST['email']
            a.status = "active"
            a.designation = request.POST['designation']
            a.password= random.SystemRandom().randint(100000, 999999)
            
            #a.branch_id = request.POST['branch']
            a.photo = request.FILES['photo']
            a.idproof = request.FILES['idproof']
            a.save()
            
            x = user_registration.objects.get(id=a.id)
            today = date.today()
            tim = today.strftime("%m%y")
            x.employee_id = "INF"+str(tim)+''+"B"+str(x.id)
            passw=x.password
            email_id=x.email
            x.save()
            y1 = user_registration.objects.get(id=a.id)
            qr = qrcode.make("http://altoscore.in/offerletter/" + str(y1.id))
            qr.save(settings.MEDIA_ROOT + "/images"+"//" +"offer"+str(y1.id) + ".png")
            with open(settings.MEDIA_ROOT + "/images"+"//"+"offer"+ str(y1.id) +".png","rb") as reopen:
                    djangofile = File(reopen)
                    y1.offerqr = djangofile
                    y1.save()
    
            y2 = user_registration.objects.get(id=a.id)
            qr1 = qrcode.make("http://altoscore.in/relieveletter/" + str(y2.id))
            qr1.save(settings.MEDIA_ROOT + "/images"+"//"+"re" +str(y2.id) + ".png")
            with open(settings.MEDIA_ROOT + "/images"+"//"+"re" + str(y2.id) + ".png", "rb") as reopen:
                    djangofile = File(reopen)
                    y2.relieveqr = djangofile
                    y2.save()
            y3 = user_registration.objects.get(id=a.id)
            qr2 = qrcode.make("http://altoscore.in/experienceletter/" + str(y3.id))
            qr2.save(settings.MEDIA_ROOT + "/images"+"//"+"exp" +str(y3.id) + ".png")
            with open(settings.MEDIA_ROOT + "/images"+"//"+"exp" + str(y3.id) + ".png", "rb") as reopen:
                    djangofile = File(reopen)
                    y3.expqr = djangofile
                    y3.save()
           
    
            b.user_id = a.id
            b.plustwo = request.POST.get('plustwo')
            b.school = request.POST['school']
            b.schoolaggregate = request.POST['aggregate']
            if request.FILES.get('cupload') is not None:
                b.schoolcertificate = request.FILES['cupload']
            b.ugdegree = request.POST['degree']
            b.ugstream = request.POST['stream']
            b.ugpassoutyr = request.POST['passoutyear']
            b.ugaggregrate = request.POST['aggregate1']
            b.backlogs = request.POST['supply']
            if request.FILES.get('cupload1') is not None:
                b.ugcertificate = request.FILES['cupload1']
            b.pg = request.POST['pg']
            b.save()
    
            c.user_id = a.id
            c.internshipdetails = request.POST['details']
            c.internshipduration = request.POST['duration']
            c.internshipcertificate = request.POST['certificate']
            c.onlinetrainingdetails = request.POST['details1']
            c.onlinetrainingduration = request.POST['duration1']
            c.onlinetrainingcertificate= request.POST['certificate1']
            c.projecttitle = request.POST['title']
            c.projectduration = request.POST['duration2']
            c.projectdescription = request.POST['description']
            c.projecturl = request.POST['url']
            c.skill1 = request.POST['skill1']
            c.skill2 = request.POST['skill2']
            c.skill3 = request.POST['skill3']
            c.save()
            
            subject = 'Greetings from ALTOS TECHNOLOGIES'
            message = 'Congratulations,\nYou have successfully registered ALTOS TECHNOLOGIES.\nYour login credentials \n\nEmail :'+str(email_id)+'\nPassword :'+str(passw)+'\n\nNote: This is a system generated email, do not reply to this email id.'
            email_from = settings.EMAIL_HOST_USER
            
            recipient_list = [email_id, ]
            send_mail(subject,message , email_from, recipient_list, fail_silently=True)
            msg_success = "Registration successfully Check Your Registered Mail"
            return redirect('login')
        
    return redirect('login')



def reset_password(request):
    if request.method == "POST":
        email_id = request.POST.get('email')
        access_user_data = user_registration.objects.filter(email=email_id).exists()
        if access_user_data:
            _user = user_registration.objects.get(email=email_id)
            password = random.SystemRandom().randint(100000, 999999)

            _user.password = password
            subject = 'iNFOX Technologies your authentication data updated'
            message = 'Password Reset Successfully\n\nYour login details are below\n\nUsername : ' + str(email_id) + '\n\nPassword : ' + str(password) + \
                '\n\nYou can login this details\n\nNote: This is a system generated email, do not reply to this email id'
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [email_id, ]
            send_mail(subject, message, email_from,
                      recipient_list, fail_silently=True)

            _user.save()
            msg_success = "Password Reset successfully check your mail new password"
            return render(request, 'home/Reset_password.html', {'msg_success': msg_success})
        else:
            msg_error = "This email does not exist iNFOX Technologies "
            return render(request, 'home/Reset_password.html', {'msg_error': msg_error})

    return render(request,'home/Reset_password.html')

def internshipform(request):
    # branch = branch_registration.objects.all()
    return render(request, 'home/internship.html')

def internship_save(request):

    a = internship()
    if request.method == 'POST':
        try:
            a.fullname = request.POST['name']
            a.collegename = request.POST['college_name']
            a.reg_date = datetime.now()
            a.reg_no = request.POST['reg_no']
            a.course = request.POST['course']
            a.stream = request.POST['stream']
            a.platform = request.POST['platform']

            a.start_date =  request.POST['start_date']
            a.end_date  =  request.POST['end_date']
            a.mobile  =  request.POST['mobile']

            a.alternative_no  =  request.POST['alternative_no']

            a.email = request.POST['email']
            a.profile_pic  =  request.FILES['profile_pic']
            if (a.end_date<a.start_date):
                return render(request,'home/internship.html',{'a':a})
            else:
                a.save()
                qr = qrcode.make("https://altoscore.in/admin_code?id=" + str(a.id))
                qr.save(settings.MEDIA_ROOT + "\\" +str(a.id) + ".png")
                with open(settings.MEDIA_ROOT + "\\" + str(a.id) + ".png", "rb") as reopen:
                        djangofile = File(reopen)
                        a.qr = djangofile

                        a.save()
           
            msg_success="Your application has been sent successfully"
            Flag='True'
            return render(request, 'home/internship.html',{'msg_success':msg_success})
        except:
            message = "Enter all details !!!"
            return render(request, 'home/internship.html',{'message':message})
    else:
        
        return render(request, 'home/internship.html')



# -----------------------------------------------------------------------------Admin Section

def ad_base(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'admin/ad_base.html',context)

def ad_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'admin/ad_profile.html',context )

def ad_dashboard(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'admin/ad_dashboard.html',context)

def ad_create_work(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'admin/ad_create_work.html',context)

def save_create_work(request):
    client = client_information()
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    if request.session.has_key('userid'):
        userid = request.session['userid']
    else:
       return redirect('/')
    if request.method == 'POST':
        client.client_name = request.POST.get('client_name')
        
        client.client_address = request.POST.get('client_address')
        client.client_mail = request.POST.get('client_mail')
        client.bs_name = request.POST.get('bs_name')
        client.bs_website = request.POST.get('bs_website',None)
       
        client.bs_location = request.POST.get('bs_location')
        client.client_files = request.FILES.get('client_files',None)
        client.seo = request.POST.get('seo',None)
        client.seo_txt = request.POST.get('seo_txt',None)
        client.seo_file = request.FILES.get('seo_file',None)
        client.seo_target = request.POST.get('seo_target',None)

        client.on_pg = request.POST.get('onpage',None)
        client.on_pg_txt = request.POST.get('on_txt',None)
        client.on_pg_file = request.FILES.get('on_file',None)
        client.on_pg_target = request.POST.get('on_target',None)

        client.off_pg = request.POST.get('offpage',None)
        client.off_pg_txt = request.POST.get('off_txt',None)
        client.off_pg_file = request.FILES.get('off_file',None)
        client.off_pg_target = request.POST.get('off_target',None)

        client.smm = request.POST.get('smm',None)
        client.smm_txt = request.POST.get('smm_txt',None)
        client.smm_file = request.FILES.get('smm_file',None)
        client.smm_target = request.POST.get('smm_target',None)
        client.smo = request.POST.get('smo',None)
        client.smo_txt = request.POST.get('smo_txt',None)
        client.smo_file = request.FILES.get('smo_file',None)
        client.smo_target = request.POST.get('smo_target',None)

        client.sem = request.POST.get('sem',None)
        client.sem_txt = request.POST.get('sem_txt',None)
        client.sem_file = request.FILES.get('sem_file',None)
        client.sem_target = request.POST.get('sem_target',None)
        client.em = request.POST.get('em',None)
        client.em_txt = request.POST.get('em_txt',None)
        client.em_file = request.FILES.get('em_file',None)
        client.em_target = request.POST.get('em_target',None)
        client.cm = request.POST.get('cm',None)
        client.cm_txt = request.POST.get('cm_txt',None)
        client.cm_file = request.FILES.get('cm_file',None)
        client.cm_target = request.POST.get('cm_target',None)
        client.am = request.POST.get('am',None)
        client.am_txt = request.POST.get('am_txt',None)
        client.am_file = request.FILES.get('am_file',None)
        client.am_target = request.POST.get('am_target',None)
        client.mm = request.POST.get('mm',None)
        client.mm_txt = request.POST.get('mm_txt',None)
        client.mm_file = request.FILES.get('mm_file',None)
        client.mm_target = request.POST.get('mm_target',None)
        client.vm = request.POST.get('vm',None)
        client.vm_txt = request.POST.get('vm_txt',None)
        client.vm_file = request.FILES.get('vm_file',None)
        client.vm_target = request.POST.get('vm_target',None)

        client.lc = request.POST.get('lc',None)
        client.lc_txt = request.POST.get('lc_txt',None)
        client.lc_file = request.FILES.get('lc_file',None)
        client.lc_target = request.POST.get('lc_target',None)

        client.user=usr
        client.save()
        
        client = client_information.objects.get(id=client.id)
        
        labels = request.POST.getlist('label[]')
        text =request.POST.getlist('dis[]')
        
        if len(labels)==len(text):
            mapped = zip(labels,text)
            mapped=list(mapped)
            for ele in mapped:
            
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
        else:
            pass

        labels2 = request.POST.getlist('label2[]')
        text2 =request.POST.getlist('dis2[]')
        
        if len(labels2)==len(text2):
            mappeds = zip(labels2,text2)
            mappeds=list(mappeds)
            for ele in mappeds:
            
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='business_details')
        else: 
            pass
          
        
        files_req =request.FILES.getlist('file_add[]') 
        label_req =request.POST.getlist('label_req[]')
        dis_req =request.POST.getlist('dis_req[]') 
        target =request.POST.getlist('target[]')

        
        if len(files_req)==len(label_req)==len(dis_req)==len(target):
            mapped2 = zip(label_req,dis_req,files_req,target)
            mapped2=list(mapped2)
         
            for ele in mapped2:
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],file=ele[2],target=ele[3],user=usr,client=client,section='requirments')

        msg_success = "Save Successfully"
        context={
            "usr":usr,
            "msg_success":msg_success,
        }
        return redirect("ad_dashboard") 
        
    return redirect("ad_create_work")


def ad_view_work(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)

    client=client_information.objects.all()

    context={
        "usr":usr,
        "client":client,
    }
    return render(request, 'admin/ad_view_work.html',context) 

def ad_view_clint(request,id):
    client=client_information.objects.get(id=id)
    addicl=addi_client_info.objects.filter(client=client.id,section='client_information')
    addibs=addi_client_info.objects.filter(client=client.id,section='business_details')
    addirq=addi_client_info.objects.filter(client=client.id,section='requirments')
    
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
        "client":client,
        "addicl":addicl,
        "addibs":addibs,
        "addirq":addirq,
    }
    return render(request, 'admin/ad_view_clint.html',context)



def update_client(request,id):
    client = client_information.objects.get(id=id)
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    if request.session.has_key('userid'):
        userid = request.session['userid']
    else:
       return redirect('/')
    if request.method == 'POST':
        client.client_name = request.POST.get('client_name')
        
        client.client_address = request.POST.get('client_address')
        client.client_mail = request.POST.get('client_mail')
        client.bs_name = request.POST.get('bs_name')
        client.bs_website = request.POST.get('bs_website',None)
       
        client.bs_location = request.POST.get('bs_location')
        if request.FILES.get('client_files',None) == None:
            client.client_files=client.client_files
        else:
            client.client_files = request.FILES.get('client_files',None)
      
        client.seo = request.POST.get('seo',None)
        client.seo_txt = request.POST.get('seo_txt',None)
        if request.FILES.get('seo_file',None) == None:
            client.seo_file=client.seo_file
        else:
            client.seo_file = request.FILES.get('seo_file',None)

        client.on_pg = request.POST.get('onpage',None)
        client.on_pg_txt = request.POST.get('on_txt',None)
        client.on_pg_target = request.POST.get('on_target',None)
        if request.FILES.get('on_file',None) == None:
            client.on_pg_file=client.on_pg_file
        else:
            client.on_pg_file = request.FILES.get('on_file',None)


        client.off_pg = request.POST.get('offpage',None)
        client.off_pg_txt = request.POST.get('off_txt',None)
        client.off_pg_target = request.POST.get('off_target',None)
        if request.FILES.get('off_file',None) == None:
            client.off_pg_file=client.off_pg_file
        else:
            client.off_pg_file = request.FILES.get('off_file',None)

   


        client.smm = request.POST.get('smm',None)
        client.smm_txt = request.POST.get('smm_txt',None)
        client.smm_target = request.POST.get('smm_target',None)
     
        if request.FILES.get('smm_file',None) == None:
            client.smm_file=client.smm_file
        else:
            client.smm_file = request.FILES.get('smm_file',None)

        client.smo = request.POST.get('smo',None)
        client.smo_txt = request.POST.get('smo_txt',None)
        client.smo_target = request.POST.get('smo_target',None)
     
        if request.FILES.get('smo_file',None) == None:
            client.smo_file=client.smo_file
        else:
            client.smo_file = request.FILES.get('smo_file',None)

        client.sem = request.POST.get('sem',None)
        client.sem_txt = request.POST.get('sem_txt',None)
        client.sem_target = request.POST.get('sem_target',None)
    

        if request.FILES.get('sem_file',None) == None:
            client.sem_file=client.sem_file
        else:
            client.sem_file = request.FILES.get('sem_file',None)


        client.em = request.POST.get('em',None)
        client.em_txt = request.POST.get('em_txt',None)
        client.em_target = request.POST.get('em_target',None)
        if request.FILES.get('em_file',None) == None:
            client.em_file=client.em_file
        else:
            client.em_file = request.FILES.get('em_file',None)


        client.cm = request.POST.get('cm',None)
        client.cm_txt = request.POST.get('cm_txt',None)
        client.cm_target = request.POST.get('cm_target',None)

        if request.FILES.get('cm_file',None) == None:
            client.cm_file=client.cm_file
        else:
            client.cm_file = request.FILES.get('cm_file',None)


        client.am = request.POST.get('am',None)
        client.am_txt = request.POST.get('am_txt',None)
        client.am_target = request.POST.get('am_target',None)
        if request.FILES.get('am_file',None) == None:
            client.am_file=client.am_file
        else:
            client.am_file = request.FILES.get('am_file',None)


        client.mm = request.POST.get('mm',None)
        client.mm_txt = request.POST.get('mm_txt',None)
        client.mm_target = request.POST.get('mm_target',None)
        if request.FILES.get('mm_file',None) == None:
            client.mm_file=client.mm_file
        else:
            client.mm_file = request.FILES.get('mm_file',None)


        client.vm = request.POST.get('vm',None)
        client.vm_txt = request.POST.get('vm_txt',None)
        client.vm_target = request.POST.get('vm_target',None)
        if request.FILES.get('vm_file',None) == None:
            client.vm_file=client.vm_file
        else:
            client.vm_file = request.FILES.get('vm_file',None)

        client.lc = request.POST.get('lc',None)
        
        client.lc_txt = request.POST.get('lc_txt',None)
     
        client.lc_target = request.POST.get('lc_target',None)
        if request.FILES.get('lc_file',None) == None:
            client.lc_file=client.lc_file
        else:
            client.lc_file = request.FILES.get('lc_file',None)


        client.user=usr
        client.save()
        client = client_information.objects.get(id=id)
       

        
        client = client_information.objects.get(id=client.id)
        
        labels = request.POST.getlist('label[]')
        text =request.POST.getlist('dis[]')
        
        if len(labels)==len(text):
            mapped = zip(labels,text)
            mapped=list(mapped)
            count = addi_client_info.objects.filter(client=id,section='client_information').count()
            lb_count=len(labels)
            print(lb_count)
            for ele in mapped:
                
                    try:
                        adiclient = addi_client_info.objects.get(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1]))
                        
                    
                        if ((adiclient.labels==ele[0]) or (adiclient.discription==ele[1])):
                            created = addi_client_info.objects.filter(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1])).update(labels=ele[0],discription=ele[1])
                        
                        
                        elif ((adiclient.labels!=ele[0]) or (adiclient.discription!=ele[1])):
                            created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
                    
                        else:
                            pass
                            
                    except:
                        created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
                




        else:
            
            pass

        labels2 = request.POST.getlist('label2[]')
        text2 =request.POST.getlist('dis2[]')
        
        if len(labels2)==len(text2):
            mappeds = zip(labels2,text2)
            mappeds=list(mappeds)

      
            for ele in mappeds:
                try:
                    adiclient=addi_client_info.objects.get(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1]))
                    if ((adiclient.labels==ele[0]) or (adiclient.discription==ele[1])):
                        created = addi_client_info.objects.filter(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1])).update(labels=ele[0],discription=ele[1])
                    elif ((adiclient.labels!=ele[0]) or (adiclient.discription!=ele[1])):
                        created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],client=client,user=usr,section='business_details')
                    else:
                        pass
                except:
                    created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],client=client,user=usr,section='business_details')

        else: 
            pass
        
      
          

        
        label_req =request.POST.getlist('label_req[]')
        dis_req =request.POST.getlist('dis_req[]')
        target =request.POST.getlist('target[]')
       
        if request.FILES.getlist('file_add[]') == []:
            img=addi_client_info.objects.filter(client=id,section='Requirments')
            files_req=[]
            for i in img:
                files_req.append(i.file)
       
           
        elif len(request.FILES.getlist('file_add[]')) != len(label_req):
          
            img=addi_client_info.objects.filter(client=id,section='Requirments')
            fr=[]
            for i in img:
                fr.append(i.file)
            fr2 =request.FILES.getlist('file_add[]') 
            files_req=fr+fr2
           
        else:
           
            files_req=request.FILES.getlist('file_add[]')

        
        
        if len(label_req)==len(dis_req)==len(target)==len(files_req):
            
            mapped2 = zip(label_req,dis_req,files_req,target)
            mapped2=list(mapped2)
           
            
            abs=addi_client_info.objects.filter(client=id,section='Requirments').delete()
            for ele in mapped2:
                    

                    created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],file=ele[2],target=ele[3],user=usr,client=client,section='Requirments')
                
        else:
       
            pass

        msg_success = "Save Successfully"
        return redirect('ad_view_clint',id)
    return redirect('ad_view_clint',id)




def ad_daily_work_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_work=daily_work.objects.filter(date=date.today())
    master=Work.objects.all()
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dl_leeds=daily_leeds.objects.all()
    print(dl_sub)
    context={
        "usr":usr,
        "dl_work":dl_work,
        "dl_leeds":dl_leeds,
        "dl_off":dl_off,
        "dl_sub":dl_sub,
        "master":master

    }
    return render(request, 'admin/ad_daily_work_det.html',context)


def ad_work_analiz_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dl_leeds=daily_leeds.objects.all()
    master=Work.objects.all()
    dl_work=daily_work.objects.all()
    context={
        "usr":usr,
        "dl_work":dl_work,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,
        "master":master

    }
    return render(request, 'admin/ad_work_analiz_det.html',context)

def flt_dt_analiz(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dl_leeds=daily_leeds.objects.all()

    dl_work=daily_work.objects.filter(date__gte=st_dt,date__lte=en_dt)
    context={
        "usr":usr,
        "dl_work":dl_work,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds

    }
    return render(request, 'admin/ad_work_analiz_det.html',context)


def ad_work_progress(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    pr_work=progress_report.objects.all()
    context={
        "usr":usr,
        "pr_work":pr_work

    }
    return render(request, 'admin/ad_work_progress.html',context)

def flt_progress(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
    
    pr_work=progress_report.objects.filter(start_date__gte=st_dt,start_date__lte=en_dt)
    context={
        "usr":usr,
        "pr_work":pr_work

    }
    return render(request, 'admin/ad_work_progress.html',context)


def ad_work_progress_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
 
    pr_work=progress_report.objects.get(id=id)
    try:
        prv_work=progress_report.objects.filter(work_id=pr_work.work_id).order_by('-end_date')[1]
    except:
        prv_work=None
    context={
        "usr":usr,
        "pr_work":pr_work,
        "prv_work":prv_work

    }
    return render(request, 'admin/ad_work_progress_det.html',context) 

# 

def ad_warning_ex(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
   
    context={
        "usr":usr,
        "exe":exe

    }
    return render(request, 'admin/ad_warning_ex.html',context)

def ad_warning_sugg_dash(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    ids=id
    context={
        "usr":usr,
        "ids":ids
        

    }
    return render(request, 'admin/ad_warning_sugg_dash.html',context)

def ad_warning_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    warn = Warning.objects.filter(executive=id,type="Warning")
    context={
        "usr":usr,
        "warn":warn

    }
    return render(request, 'admin/ad_warning_det.html',context) 

def ad_suggestions_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    warn = Warning.objects.filter(executive=id,type="Suggestion")
    context={
        "usr":usr,
        "warn":warn

    }
    return render(request, 'admin/ad_suggestions_det.html',context)


def change_pass(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)

    if request.method == 'POST':
        abc = user_registration.objects.get(id=devid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'admin/ch_pass.html', {'dev': dev,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'admin/ch_pass.html', {'dev': dev,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'admin/ch_pass.html', {'dev': dev,"usr":usr})
    return render(request, 'admin/ch_pass.html', {'dev': dev,"usr":usr})

def ad_accountset(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    return render(request, 'admin/ad_accountset.html', {'dev': dev,"usr":usr})

def ad_imagechange(request, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    if request.method == 'POST':
        abc = user_registration.objects.get(id=id)
        abc.photo = request.FILES['filename']
        
        abc.save()
        return redirect('ad_accountset')
    return render(request, 'admin/ad_accountset.html',{'dev': dev,"usr":usr})


def get_dis(request):
    ele = request.GET.get('ele')
    warn = daily_work.objects.get(id=ele)
    
    rep =warn.workdone
 
    return JsonResponse({"status":" not","rep":rep})

def get_sub(request):
    ele = request.GET.get('ele')
    ids = request.GET.get('idss')
    try:
        warn = daily_work.objects.get(id=ids)
    except:
        pass
    if ele=="Facebook":
        hd=ele
        des=warn.fb_txt
        fl=warn.fb_file

    elif ele=="Twitter":
        hd=ele
        des=warn.tw_txt
        fl=warn.tw_file

    elif ele=="Pinterest":
        hd=ele
        des=warn.pin_txt
        fl=warn.pin_file

    elif ele=="Linkedin":
        hd=ele
        des=warn.link_txt
        fl=warn.link_file

    elif ele=="Instagram":
        hd=ele
        des=warn.insta_txt
        fl=warn.insta_file

    elif ele=="Tumblr":
        hd=ele
        des=warn.tumb_txt
        fl=warn.tumb_file

    elif ele=="Directories":
        hd=ele
        des=warn.diry_txt
        fl=warn.diry_file

    elif ele=="You Tube":
        hd=ele
        des=warn.yt_txt
        fl=warn.yt_file

    elif ele=="Quora":
        hd=ele
        des=warn.qra_txt
        fl=warn.qra_file

    elif ele=="PR Submission":
        hd=ele
        des=warn.pr_txt
        fl=warn.pr_file 

    elif ele=="Article Submission":
        hd=ele
        des=warn.art_txt
        fl=warn.art_file 

    elif ele=="Blog Posting":
        hd=ele
        des=warn.blg_txt
        fl=warn.blg_file 

    elif ele=="Classified Submission":
        hd=ele
        des=warn.clss_txt
        fl=warn.clss_file

    elif ele=="Guest Blogging":
        hd=ele
        des=warn.gst_txt
        fl=warn.gst_file

    elif ele=="Bokkmarking":
        hd=ele
        des=warn.bk_txt
        fl=warn.bk_file

    elif daily_off_sub.objects.filter(id=ids,sub=ele).exists():
       
        off = daily_off_sub.objects.get(id=ids,sub=ele)
        hd=off.sub
        des=off.sub_txt
        fl=off.sub_file

    elif daily_leeds.objects.filter(id=ids,sub=ele).exists():
       
        lc = daily_leeds.objects.get(id=ids,sub=ele)
        hd=lc.sub
        des=lc.sub_txt
        fl=lc.sub_file


    else:
        
        sm = daily_work_sub.objects.get(id=ids,sub=ele)
        hd=sm.sub
        des=sm.sub_txt
        fl=sm.sub_file
        

    return JsonResponse({"status":" not","hd":hd,"des":des,"fl":str(fl),})



def work_shedule_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
   
    context={
        "usr":usr,
        "exe":exe

    }
    return render(request, 'admin/ad_exe_shedule.html',context)

def work_shedule(request,id):
    ids=request.session['userid']
    now = datetime.now()
    usr = user_registration.objects.get(id=ids)
    events = Events.objects.filter(executive=id, start__date=now.date())
    dl_sub=addi_events.objects.filter(executive=id)

    

    # # Define a start and end time range
    # start_time = time(0, 0, 0)
    # end_time = time(23, 59, 59)

    # # Define the date for which you want to find missing times
    # search_date = datetime(2023, 5, 4).date()

    # # Create a list of all times that should be in the database for the given date
    # expected_times = [datetime.combine(search_date, start_time) + timedelta(seconds=x) for x in range((datetime.combine(search_date, end_time) - datetime.combine(search_date, start_time)).seconds + 1)]

    # # Retrieve all date and time values from the database for the given date
    # actual_datetimes = Events.objects.filter(start__date=search_date).values_list('start', flat=True)

    # # Extract the time portion of the actual date and time values
    # actual_times = [dt.time() for dt in actual_datetimes]

    # # Find missing times
    # missing_times = set(expected_times) - set(actual_times)

    # # Print out the missing times
    # print(missing_times)

    # from datetime import datetime, timedelta, time
    # from django.db.models.functions import TruncHour
  
    # start_time = time(0, 0, 0)
    # end_time = time(9, 59, 59)

    # search_date = datetime(2023, 5, 4).date()


    # expected_hours = [datetime.combine(search_date, start_time) + timedelta(hours=x) for x in range((datetime.combine(search_date, end_time) - datetime.combine(search_date, start_time)).seconds // 3600 + 1)]

  
    # actual_hours = Events.objects.filter(start__date=search_date).annotate(hour=TruncHour('start')).values_list('start', flat=True)

    # # Find missing hours
    # missing_hours = set(expected_hours) - set(actual_hours)

    # # Print out the missing hours
    # print(missing_hours)


    # from datetime import datetime, timedelta, time
    # from django.db.models.functions import TruncHour


    # # Define a start and end date and time range
    # start_datetime = datetime(2023, 4, 5, 9, 0, 0)
    # end_datetime = datetime(2023, 4, 5, 18, 0, 0)

    # # Create a list of all hours that should be in the table for the given date range
    # expected_hours = []
    # current_datetime = start_datetime
    # while current_datetime <= end_datetime:
    #     expected_hours += [current_datetime]
    #     current_datetime += timedelta(hours=1)

    # # Retrieve all records from the database for the given date range and group them by hour
    # actual_hours = Events.objects.filter(start__gte=start_datetime, end__lte=end_datetime, id=1).annotate(hour=TruncHour('start')).values_list('hour', flat=True)
   
    # # Find missing hours
    # missing_hours = set(expected_hours) - set(actual_hours)

    # # Print out the missing hours
    # print(missing_hours)

    
    


    

    start_time = datetime.combine(now.date(), time.min)
    end_time = datetime.combine(now.date(), time.max)
    records = Events.objects.filter(start__date=now.date(), end__date=now.date(), start__lte=end_time, end__gte=start_time,executive=id)
    # records = Events.objects.filter(start__lte=start_time, end__gte=end_time)

    all_hours = [start_time + timedelta(hours=x) for x in range(9,18)]
    
    occupied_hours = set()
    for record in records:
        
        hours = [record.start+ timedelta(hours=x) for x in range((record.end - record.start).seconds // 3600 + 1)]
        occupied_hours.update(hours)
   
 

    free_hours = set(all_hours) - occupied_hours
    

    lst = []
    
    
    # for all_hr in sorted(all_hours):
    #   for oc_hr in sorted(occupied_hours):
    #         print( str(all_hr.time()) +"="+ str(oc_hr.time()))
    #         if all_hr.time() == oc_hr.time():
    #             pass
    #         else:
                
    #             print(all_hr.time())
    
    
 
    list1=[]
    for ls1 in sorted(occupied_hours):
        list1.append(ls1.time())

    list2=[]
    for ls2 in all_hours:
        list2.append(ls2.time())

    result_set = set(list1) - set(list2)

    result_list = list(result_set)
    
    def remove_dupilicate(List1,List2):
        return [item for item in List1 if item not in List2]

    new_a = remove_dupilicate(list1, list2)
    new_b = remove_dupilicate(list2, list1)

    k=list(sorted(new_b))
    
    context={
        "usr":usr,
        'events': events,
        "hr":k,
        "noti":len(k),
        "idr":id,
        "dl_sub":dl_sub
        

    }
    return render(request, 'admin/ad_work_shedule.html',context)

def filter_shedule(request,id):

    if request.method == 'POST':
        dates = request.POST['flt_date']
        
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        events = Events.objects.filter(executive=id, start__date=dates)
        dl_sub=addi_events.objects.filter(executive=id)
     
        now = datetime.now()

        start_time = datetime.combine(now.date(), time.min)
        end_time = datetime.combine(now.date(), time.max)
        
        records = Events.objects.filter(start__date=dates, end__date=dates,executive=id)
     

        all_hours = [start_time + timedelta(hours=x) for x in range(9,18)]
        
        occupied_hours = set()
        for record in records:
            
            hours = [record.start+ timedelta(hours=x) for x in range((record.end - record.start).seconds // 3600 + 1)]
            occupied_hours.update(hours)
    
    

        free_hours = set(all_hours) - occupied_hours
        

        lst = []
        
    
        list1=[]
        for ls1 in sorted(occupied_hours):
            list1.append(ls1.time())

        list2=[]
        for ls2 in all_hours:
            list2.append(ls2.time())

        result_set = set(list1) - set(list2)

        result_list = list(result_set)
        
        def remove_dupilicate(List1,List2):
            return [item for item in List1 if item not in List2]

        new_a = remove_dupilicate(list1, list2)
        new_b = remove_dupilicate(list2, list1)

        k=list(sorted(new_b))
        
        context={
            "usr":usr,
            'events': events,
            "hr":k,
            "noti":len(k),
            "idr":id,
            "dl_sub":dl_sub

        }
        return render(request, 'admin/ad_work_shedule.html',context)
    return redirect("work_shedule",id)

# views.py


def events(request):
    events = Events.objects.all()
    data = []
    for event in events:
        data.append({
            'title': event.title,
            'start': event.start_time.isoformat(),
            'end': event.end_time.isoformat(),
        })
    return JsonResponse(data, safe=False)

def ad_exe_smopost(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
   
    context={
        "usr":usr,
        "exe":exe

    }
    return render(request, 'admin/ad_exe_smopost.html',context) 
    # 

def ad_sv_smopost(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    post = smo_post.objects.filter(executive=id)
    addi_post=addi_smo_post.objects.all()
    print(addi_post)
    context={
            "usr":usr,
            "post":post,
            "addi_post":addi_post,
        }
    return render(request, 'admin/ad_sv_smopost.html',context)

def ad_get_smo_pst(request):
    ele = request.GET.get('ele')
    ids = request.GET.get('idss')
    print("fdgfdgfdgdgfdgfdg")
    print(ids)
    try:
        warn = smo_post.objects.get(id=ids)
    except:
        pass
    if ele=="Facebook":
        print("haii")
        hd=ele
        des=warn.fb_dt
        print(des)
        fl=warn.fb_file

    elif ele=="Twitter":
        hd=ele
        des=warn.tw_dt
        fl=warn.tw_file

    elif ele=="Pinterest":
        hd=ele
        des=warn.pin_dt
        fl=warn.pin_file

    elif ele=="Linkedin":
        hd=ele
        des=warn.link_dt
        fl=warn.link_file

    elif ele=="Instagram":
        hd=ele
        des=warn.insta_dt
        fl=warn.insta_file

    elif ele=="Tumblr":
        hd=ele
        des=warn.tumb_dt
        fl=warn.tumb_file

    elif ele=="Directories":
        hd=ele
        des=warn.diry_dt
        fl=warn.diry_file

    elif ele=="You Tube":
        hd=ele
        des=warn.yt_dt
        fl=warn.yt_file

    elif ele=="Quora":
        hd=ele
        des=warn.qra_dt
        fl=warn.qra_file

    elif ele=="PR Submission":
        hd=ele
        des=warn.pr_dt
        fl=warn.pr_file 

    elif ele=="Article Submission":
        hd=ele
        des=warn.art_dt
        fl=warn.art_file 

    elif ele=="Blog Posting":
        hd=ele
        des=warn.blg_dt
        fl=warn.blg_file 

    elif ele=="Classified Submission":
        hd=ele
        des=warn.clss_dt
        fl=warn.clss_file

    elif ele=="Guest Blogging":
        hd=ele
        des=warn.gst_dt
        fl=warn.gst_file

    elif ele=="Bokkmarking":
        hd=ele
        des=warn.bk_dt
        fl=warn.bk_file

    elif daily_off_sub.objects.filter(id=ids,sub=ele).exists():
       
        off = daily_off_sub.objects.get(id=ids,sub=ele)
        hd=off.sub
        des=off.sub_dt
        fl=off.sub_file

    elif addi_smo_post.objects.filter(id=ids,label=ele).exists():
      
        add_smo = addi_smo_post.objects.get(id=ids,label=ele)
        hd=add_smo.label
        des=add_smo.date
        fl=add_smo.file
    else:
        
        sm = daily_work_sub.objects.get(id=ids,sub=ele)
        hd=sm.sub
        des=sm.sub_dt
        fl=sm.sub_file
        

    return JsonResponse({"status":" not","hd":hd,"des":des,"fl":str(fl),})

def get_event_det(request):
    ele = request.GET.get('ele')
    ids = request.GET.get('idss')
 
    try:
        warn = Events.objects.get(id=ids)
    except:
        warn = addi_events.objects.get(id=ids)
    if ele=="Facebook":
        print("haii")
        hd=ele
        des=warn.fb_dt
        print(des)
        fl=warn.fb_file

    elif ele=="Twitter":
        hd=ele
        des=warn.tw_dt
        fl=warn.tw_file

    elif ele=="Pinterest":
        hd=ele
        des=warn.pin_dt
        fl=warn.pin_file

    elif ele=="Linkedin":
        hd=ele
        des=warn.link_dt
        fl=warn.link_file

    elif ele=="Instagram":
        hd=ele
        des=warn.insta_dt
        fl=warn.insta_file

    elif ele=="Tumblr":
        hd=ele
        des=warn.tumb_dt
        fl=warn.tumb_file

    elif ele=="Directories":
        hd=ele
        des=warn.diry_dt
        fl=warn.diry_file

    elif ele=="You Tube":
        hd=ele
        des=warn.yt_dt
        fl=warn.yt_file

    elif ele=="Quora":
        hd=ele
        des=warn.qra_dt
        fl=warn.qra_file

    elif ele=="PR Submission":
        hd=ele
        des=warn.pr_dt
        fl=warn.pr_file 

    elif ele=="Article Submission":
        hd=ele
        des=warn.art_dt
        fl=warn.art_file 

    elif ele=="Blog Posting":
        hd=ele
        des=warn.blg_dt
        fl=warn.blg_file 

    elif ele=="Classified Submission":
        hd=ele
        des=warn.clss_dt
        fl=warn.clss_file

    elif ele=="Guest Blogging":
        hd=ele
        des=warn.gst_dt
        fl=warn.gst_file

    elif ele=="Bokkmarking":
        hd=ele
        des=warn.bk_dt
        fl=warn.bk_file

    elif addi_events.objects.filter(id=ids,label=ele).exists():
       
        off = addi_events.objects.get(id=ids,label=ele)
        hd=off.label
        des=off.date
        fl=off.file
    else:
        
        sm = daily_work_sub.objects.get(id=ids,sub=ele)
        hd=sm.sub
        des=sm.sub_dt
        fl=sm.sub_file
        

    return JsonResponse({"status":" not","hd":hd,"des":des,"fl":str(fl),})


def ad_export_excel(request,id):

    filtered_data = daily_leeds.objects.filter(daily=id)

    # Create an Excel workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Add column headers to the Excel sheet
    headers = ['No.',"name","email_id","ph_no","location","qualification","year_of_passout","collegename","internship","internship_institute","internship_topic","internship_start","internship_end","duration","fresher_experience","previous_experience","company_name","register for what",'Duration' 'Experiense','Internship']  # "Replace with your actual column names
    sheet.append(headers)

    # Add data rows to the Excel sheet
    count = 1
    for item in filtered_data:
        
        row = [count,item.name,item.email_id,item.ph_no,item.location,item.qualification,item.year_of_passout,item.collegename,item.internship,item.internship_institute,item.internship_topic,item.internship_start,item.internship_end,item.duration,item.fresher_experience,item.previous_experience,item.company_name,item.register,item.ex_duration] # Replace with your actual column names
        sheet.append(row)
        count+=1

    # Set the response headers for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filtered_data.xlsx'

    # Save the Excel workbook to the response
    workbook.save(response)

    return response

def ad_leave_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    context={
        "usr":usr,
      
    }
    return render(request, 'admin/ad_leave_home.html',context)

def ad_tl_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Digital Marketing Head")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/ad_tl_det.html',context)

def view_tl_leave(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=leave.objects.filter(user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id
      
    }
    return render(request, 'admin/ad_view_tl_leave.html',context)

def flt_leave_tl(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
    
    tls=leave.objects.filter(from_date__gte=st_dt,from_date__lte=en_dt,user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id

    }
    return render(request, 'admin/ad_view_tl_leave.html',context)


def ad_exe_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/ad_exe_det.html',context)



def view_exe_leave(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=leave.objects.filter(user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id
      
    }
    return render(request, 'admin/ad_view_exe_leave.html',context)

def flt_leave_exe(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
    
    tls=leave.objects.filter(from_date__gte=st_dt,from_date__lte=en_dt,user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id

    }
    return render(request, 'admin/ad_view_exe_leave.html',context)

def ad_tele_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Telecaller")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/ad_tele_det.html',context)

def view_tele_leave(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=leave.objects.filter(user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id
      
    }
    return render(request, 'admin/ad_view_tele_leave.html',context)

def view_tele_leave(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=leave.objects.filter(user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id
      
    }
    return render(request, 'admin/ad_view_tele_leave.html',context)

def flt_leave_tele(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
    
    tls=leave.objects.filter(from_date__gte=st_dt,from_date__lte=en_dt,user=id)
    context={
        "usr":usr,
        "tls":tls,
        "id":id

    }
    return render(request, 'admin/ad_view_tele_leave.html',context)


def all_leads_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/ad_all_leads_exe.html',context)

def ad_view_exe_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads_by_date = All_leads.objects.filter(executive=id).values('date').annotate(total_leads=Count('id'))
    for lead in leads_by_date:
        lead['target'] = random.randint(70, 120)
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        "id":id
    }
    return render(request, 'admin/ad_view_exe_det.html', context)

def ad_view_all_leads(request, date, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if date:
        leads = All_leads.objects.filter(date=date, executive=id)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        "id":id,
        "date":date
    }
    return render(request, 'admin/ad_view_all_leads.html', context)

def ad_filter_day_previous_leads(request,date,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    
    if date:
        leads = All_leads.objects.filter(date=day_field, executive=id)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        "id":id,
        "date":date
    }
    return render(request, 'admin/ad_view_all_leads.html', context)


def ad_filter_month_previous_leads(request,date,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

   

    if date:
        leads = All_leads.objects.filter(Q(date__year=selected_year_number) & Q(date__month=selected_month_number), executive=id)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        "id":id,
        "date":date
    }
    return render(request, 'admin/ad_view_all_leads.html', context)


def ad_assign_Leads(request):
    
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assign=All_leads.objects.all().values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')
    telecaller=user_registration.objects.filter(department='Telecaller')
    print(assign)
    context={
        "usr":usr,
        'assign':assign,
        'telecaller':telecaller,       

    }
    return render(request,'admin/ad_assigned_tele.html', context)

def ad_assigned_person_details(request,id):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assign=All_leads.objects.filter(telecaller_id=id).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'assign':assign,
        "id":id


    }
    return render(request,'admin/ad_assign_tele_det.html', context)

def ad_filter_day_previous_assign(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    
    if date:
        assign=All_leads.objects.filter(assign_dt=day_field, telecaller_id=id)
    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,

    }
    return render(request,'admin/ad_assign_tele_det.html', context)


def ad_filter_month_previous_assign(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

   

    if date:
        assign=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number), telecaller_id=id)
    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,
       
    }
    return render(request,'admin/ad_assign_tele_det.html', context)

def all_follow_tele(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Telecaller")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/ad_follow_tele.html',context)

def ad_follow_tele_datewise(request,tid):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    telecaller = user_registration.objects.get(id=tid)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']

    leads_by_date = All_leads.objects.filter(telecaller_id=tid,status__in=status_values).values('assign_dt').annotate(total_leads=Count('id'))

    
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        'tid': tid
    }
    return render(request, 'admin/ad_follow_tele_datewise.html', context)

def ad_follow_tele_det(request,tid,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']
    assign=All_leads.objects.filter(status__in=status_values,telecaller_id=tid,assign_dt=date).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'assign':assign,
        "tid":tid


    }
    return render(request,'admin/ad_follow_tele_det.html', context)

def ad_follow_tele_det_day(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']
    if date:
        assign=All_leads.objects.filter(status__in=status_values,assign_dt=day_field, telecaller_id=id)
    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,

    }
    return render(request,'admin/ad_follow_tele_det.html', context)


def ad_follow_tele_det_month(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']

    if date:
        assign=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number), telecaller_id=id,status__in=status_values)
    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,
       
    }
    return render(request,'admin/ad_follow_tele_det.html', context)


def all_nocall_tele(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Telecaller")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'admin/all_nocall_tele.html',context)


def ad_nocall_tele_det(request,id):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['no',""]
    # assign=All_leads.objects.filter(status__in=status_values,telecaller_id=id).order_by('-assign_dt')
    assign=All_leads.objects.filter(telecaller_id=id,status__in=status_values).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

   
    context={
        "usr":usr,
        'assign':assign,
        "id":id


    }
    return render(request,'admin/ad_nocall_tele_det.html', context)

def ad_nocall_tele_det_day(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    status_values=['no',""]
    if date:
        # assign=All_leads.objects.filter(status__in=status_values,assign_dt=day_field, telecaller_id=id)
        assign=All_leads.objects.filter(telecaller_id=id,status__in=status_values,assign_dt=day_field).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,

    }
    return render(request,'admin/ad_nocall_tele_det.html', context)


def ad_nocall_tele_det_month(request,id):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

    status_values=['no',""]

    if date:
        assign=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number), telecaller_id=id,status__in=status_values).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    else:
        assign=All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'assign': assign,
        "id":id,
       
    }
    return render(request,'admin/ad_nocall_tele_det.html', context)

def ad_nocall_tele_details(request,id,pk):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['no',""]
    # assign=All_leads.objects.filter(status__in=status_values,telecaller_id=id).order_by('-assign_dt')
    assign=All_leads.objects.filter(telecaller_id=id,status__in=status_values,assign_dt=pk).order_by('-assign_dt')

   
    context={
        "usr":usr,
        'assign':assign,
        "id":id


    }
    return render(request,'admin/ad_nocall_tele_details.html', context)


def ad_delay_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_task=lead_delay.objects.all()
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "dl_task":dl_task,
        "exe":exe
    }
    return render(request, 'admin/ad_delay_det.html', context)


def ad_delay_flt(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    if request.method=="POST":
        exe=request.POST.get('exe_name')
        start_date=request.POST.get('start_date')
        end_date=request.POST.get('end_date')
        w_mails = dm_warning_mails.objects.filter(executive_id=exe)

        st_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        # dl_task = lead_delay.objects.filter(date__gte=st_dt, date__lte=end_dt,executive_id=exe).order_by('-date')

        wk_assign = work_asign.objects.filter(exe_name_id=exe)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,start_date__gte=st_dt,end_date__lte=end_dt,status__in=status_values).order_by('-start_date')
        print(work)
        for i in work:
            if i.target is not None:
                i.difference = i.target - 10
            else:
                i.difference = None 

        today = date.today()
        context = {
            "usr":usr,
            # 'dl_task': dl_task,
            "exe":exes,
            # 'warning':warning,
            'eid':exe,
            'work':work,
            'today':today,
            'w_mails':w_mails

        }
    
        return render(request,'admin/ad_delay_flt.html', context)
    return redirect("ad_delay")



# -----------------------------------------------------------------------------Executive Section

def ex_base(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'executive/ex_base.html',context)

def ex_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work_as=work_asign.objects.filter(exe_name=ids)
    work=Work.objects.filter(start_date=date.today())
    cl=client_information.objects.all()
    last=work_asign.objects.filter(exe_name=ids).last()

    curent_month= date.today().month
    curret_year= date.today().year

    context={
        "usr":usr,
        "work":work,
        "work_as":work_as,
        "cl":cl,
        "last":last,
    }
    
    return render(request, 'executive/ex_profile.html',context)

def ex_dashboard(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    cor=correction.objects.filter(executive=ids).order_by('id').last()
    dt=date.today()
    
    dt=date.today()
    context={
        "usr":usr,
        "cor":cor,
        "dt":dt,
    }
    return render(request, 'executive/ex_dashboard.html',context)

def ex_daily_work_clint(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work_as=work_asign.objects.filter(exe_name=ids).values('client_name_id').distinct()
    last=work_asign.objects.filter(exe_name=ids).last()
    work=Work.objects.all()
    cl=client_information.objects.all()



    print(work.values)
    context={
        "usr":usr,
        "work_as":work_as,
        "work":work,
        "cl":cl,
        "last":last
    }
    return render(request, 'executive/ex_daily_work_clint.html',context)

def ex_daily_work_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work_as=work_asign.objects.filter(exe_name=ids)
    works=Work.objects.filter(client_name_id=id).order_by("-id")
    daily=daily_work.objects.filter(user=ids)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dts=date.today()
    dl_leeds=daily_leeds.objects.all()
    cr_date=date.today()
    count=daily_leeds_exists.objects.filter(date=dts,executive=usr).count()

    context={
        "usr":usr,
        "cr_date":cr_date,
        "daily":daily,
        "work_as":work_as,
        "works":works,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,
        "dts":dts,
        "count": count
        
    }
    return render(request, 'executive/ex_daily_work_det.html',context)

def daily_work_done(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    work=Work.objects.get(id=id)
    if request.method == 'POST':
        daily = daily_work()
        daily.task=work.task
        daily.date=date.today()
        daily.workdone =request.POST.get('workdone',None)

        daily.fb = request.POST.get('fb',None)
        daily.fb_txt = request.POST.get('fb_txt',None)
        daily.fb_file = request.FILES.get('fb_file',None)
        daily.tw = request.POST.get('tw',None)
        daily.tw_txt = request.POST.get('tw_txt',None)
        daily.tw_file = request.FILES.get('tw_file',None)
        daily.pin = request.POST.get('pin',None)
        daily.pin_txt = request.POST.get('pin_txt',None)
        daily.pin_file = request.FILES.get('pin_file',None)
        daily.link = request.POST.get('link',None)
        daily.link_txt = request.POST.get('link_txt',None)
        daily.link_file = request.FILES.get('link_file',None)
        daily.insta = request.POST.get('insta',None)
        daily.insta_txt = request.POST.get('insta_txt',None)
        daily.insta_file = request.FILES.get('insta_file',None)
        daily.tumb = request.POST.get('tumb',None)
        daily.tumb_txt = request.POST.get('tumb_txt',None)
        daily.tumb_file = request.FILES.get('tumb_file',None)
        daily.diry = request.POST.get('diry',None)
        daily.diry_txt = request.POST.get('diry_txt',None)
        daily.diry_file = request.FILES.get('diry_file',None)
        daily.yt = request.POST.get('yt',None)
        daily.yt_txt = request.POST.get('yt_txt',None)
        daily.yt_file = request.FILES.get('yt_file',None)
        daily.qra = request.POST.get('qra',None)
        daily.qra_txt = request.POST.get('qra_txt',None)
        daily.qra_file = request.FILES.get('qra_file',None)
        daily.sbms = request.POST.get('sbms',None)
        daily.sbms_txt = request.POST.get('sbms_txt',None)
        daily.sbms_file = request.FILES.get('sbms_file',None)

        daily.pr = request.POST.get('pr',None)
        daily.pr_txt = request.POST.get('pr_txt',None)
        daily.pr_file = request.FILES.get('pr_file',None)
        daily.art = request.POST.get('art',None)
        daily.art_txt = request.POST.get('art_txt',None)
        daily.art_file = request.FILES.get('art_file',None)
        daily.blg = request.POST.get('blg',None)
        daily.blg_txt = request.POST.get('blg_txt',None)
        daily.blg_file = request.FILES.get('blg_file',None)
        daily.clss = request.POST.get('cls',None)
        daily.clss_txt = request.POST.get('cls_txt',None)
        daily.clss_file = request.FILES.get('cls_file',None)
        daily.gst = request.POST.get('gst',None)
        daily.gst_txt = request.POST.get('gst_txt',None)
        daily.gst_file = request.FILES.get('gst_file',None)
        daily.bk = request.POST.get('bk',None)
        daily.bk_txt = request.POST.get('bk_txt',None)
        daily.bk_file = request.FILES.get('bk_file',None)
        
        dct_file = dict(request.FILES)
        lst_screenshot = dct_file['filed']
        lst_file = []
        for ins_screenshot in lst_screenshot:
            str_img_path = ""
            if ins_screenshot:
                img_emp = ins_screenshot
                fs = FileSystemStorage(location=settings.MEDIA_ROOT,base_url=settings.MEDIA_URL)
                str_img = fs.save(''.join(filter(str.isalnum, str(img_emp))), img_emp)
                str_img_path = fs.url(''.join(filter(str.isalnum, str_img)))
                lst_file.append('/media/'+''.join(filter(str.isalnum, str(img_emp))))
                daily.json_testerscreenshot = lst_file
        daily.work=work
        daily.user=usr
        daily.cl_name=work.cl_name
        daily.save() 
        dl = daily_work.objects.get(id=daily.id)
        
        sub_lb =request.POST.getlist('sub_lb[]') 
        sub_txt =request.POST.getlist('sub_txt[]')
        sub_file =request.FILES.getlist('sub_file[]')
        
        if len(sub_lb)==len(sub_txt)==len(sub_file):
            mapped2 = zip(sub_lb,sub_txt,sub_file)
            mapped2=list(mapped2)
            for ele in mapped2:
               
                created = daily_work_sub.objects.get_or_create(sub=ele[0],sub_txt=ele[1],sub_file=ele[2],daily=dl)
                    
        off_sub_lb =request.POST.getlist('off_sub_lb[]') 
        off_sub_txt =request.POST.getlist('off_sub_txt[]')
        off_sub_file =request.FILES.getlist('off_sub_file[]')
        
        if len(off_sub_lb)==len(off_sub_txt)==len(off_sub_file):
            mapped2 = zip(off_sub_lb,off_sub_txt,off_sub_file)
            mapped2=list(mapped2)
            for ele in mapped2:
               
                created = daily_off_sub.objects.get_or_create(sub=ele[0],sub_txt=ele[1],sub_file=ele[2],daily=dl)

       

        name=request.POST.getlist('lc_name[]')
        email_id=request.POST.getlist('lc_email_id[]')
        ph_no=request.POST.getlist('lc_ph_no[]')
        location=request.POST.getlist('lc_location[]')
        qualification=request.POST.getlist('lc_qualification[]')
        year_of_passout=request.POST.getlist('lc_year_of_passout[]')
        collegename=request.POST.getlist('lc_collegename[]')
        internship=request.POST.getlist('lc_internship[]')
        internship_institute=request.POST.getlist('lc_internship_institute[]')
        internship_topic=request.POST.getlist('lc_internship_topic[]')
        internship_start=request.POST.getlist('lc_internship_start[]')
        internship_end=request.POST.getlist('lc_internship_end[]')
        fresher_experience=request.POST.getlist('lc_fresher_experience[]')
        previous_experience=request.POST.getlist('lc_previous_experience[]')
        company_name=request.POST.getlist('lc_company_name[]')
        register=request.POST.getlist('lc_register[]')
        duration=request.POST.getlist('lc_duration[]')
        ex_duration=request.POST.getlist('ex_duration[]')
        print("ex_duration")
        print(ex_duration)
        if len(name)==len(email_id)==len(ph_no)==len(location)==len(qualification)==len(year_of_passout)==len(collegename)==len(internship)==len(internship_institute)==len(internship_topic)==len(internship_start)==len(internship_end)==len(fresher_experience)==len(previous_experience)==len(company_name)==len(register)==len(duration)==len(ex_duration):
            mapped2 = zip(name,email_id,ph_no,location,qualification,year_of_passout,collegename,internship,internship_institute,internship_topic,internship_start,internship_end,fresher_experience,previous_experience,company_name,register,duration,ex_duration)
            mapped2=list(mapped2)
            for ele in mapped2:
                tp=int(ele[17])
                print(type(tp))
                print("ele[17]")
                if daily_leeds.objects.filter(name=ele[0],email_id=ele[1],ph_no=ele[2]).exists():
                    if str(ele[7]) == "no":
                        if str(ele[12]) == "fresher":
                            created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=None,daily=dl,ex_duration=None,executive=usr)
                        else:
                            created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=None,daily=dl,ex_duration=ele[17],executive=usr)

                    else:
                        if str(ele[12]) == "fresher":
                            fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                            todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                            created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=None,company_name=None,register=None,duration=ele[16],daily=dl,ex_duration=None,executive=usr)
                        else:
                            fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                            todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                            created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=ele[16],daily=dl,ex_duration=ele[17],executive=usr)
                else:
                    if int(ele[16]) > 6:
                        if str(ele[7]) == "no":
                            if str(ele[12]) == "fresher":
                                created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=None,daily=dl,ex_duration=None,executive=usr)
                            else:
                                created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=None,daily=dl,ex_duration=ele[17],executive=usr)

                        else:
                            if str(ele[12]) == "fresher":
                                fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=None,company_name=None,register=None,duration=ele[16],daily=dl,ex_duration=None,executive=usr)
                            else:
                                fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=ele[16],daily=dl,ex_duration=ele[17],executive=usr)
                    else:
                        
                        if int(ele[17]) > 12:
                            if str(ele[7]) == "no":
                                if str(ele[12]) == "fresher":
                                    created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=None,daily=dl,ex_duration=None,executive=usr)
                                else:
                                    created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=None,daily=dl,ex_duration=ele[17],executive=usr)

                            else:
                                if str(ele[12]) == "fresher":
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=None,company_name=None,register=None,duration=ele[16],daily=dl,ex_duration=None,executive=usr)
                                else:
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = daily_leeds_exists.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=ele[16],daily=dl,ex_duration=ele[17],executive=usr)
                        else:
                            
                            if str(ele[7]) == "no":
                                if str(ele[12]) == "fresher":
                                    created = daily_leeds.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=None,daily=dl,ex_duration=None)
                                else:
                                    created = daily_leeds.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=None,daily=dl,ex_duration=ele[17])

                            else:
                                if str(ele[12]) == "fresher":
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = daily_leeds.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=None,company_name=None,register=None,duration=ele[16],daily=dl,ex_duration=None)
                                else:
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = daily_leeds.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=ele[16],daily=dl,ex_duration=ele[17])

                    
                    # fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                    # todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                    # created = daily_leeds.objects.get_or_create(name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],duration=ele[16],register=ele[15],daily=dl)

                    

        if len(name)==len(email_id)==len(ph_no)==len(location)==len(qualification)==len(year_of_passout)==len(collegename)==len(internship)==len(internship_institute)==len(internship_topic)==len(internship_start)==len(internship_end)==len(fresher_experience)==len(previous_experience)==len(company_name)==len(register)==len(duration)==len(ex_duration):
            mapped2 = zip(name,email_id,ph_no,location,qualification,year_of_passout,collegename,internship,internship_institute,internship_topic,internship_start,internship_end,fresher_experience,previous_experience,company_name,register,duration,ex_duration)
            mapped2=list(mapped2)
            for ele in mapped2:
                
                if All_leads.objects.filter(name=ele[0],email_id=ele[1],ph_no=ele[2]).exists():
                    
                    pass
                else:
                    print(ele[16])
                    if int(ele[16]) > 6:
                        pass
                    else:
                        
                        if int(ele[17]) > 12:
                            pass
                        else:
                            print(ele[16])
                            if str(ele[7]) == "no":
                                if str(ele[12]) == "fresher":
                                    created = All_leads.objects.get_or_create(date=date.today(),name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=None,ex_duration=None,executive=usr)
                                else:
                                    created = All_leads.objects.get_or_create(date=date.today(),name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=None,ex_duration=ele[17],executive=usr)

                            else:
                                if str(ele[12]) == "fresher":
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = All_leads.objects.get_or_create(date=date.today(),name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=None,company_name=None,register=ele[15],duration=ele[16],executive=usr)
                                else:
                                    fromdate=datetime.strptime(str(ele[10]), "%Y-%m-%d").date()
                                    todate=datetime.strptime(str(ele[11]), "%Y-%m-%d").date()
                                    created = All_leads.objects.get_or_create(date=date.today(),name=ele[0],email_id=ele[1],ph_no=ele[2],location=ele[3],qualification=ele[4],year_of_passout=ele[5],collegename=ele[6],internship=ele[7],internship_institute=ele[8],internship_topic=ele[9],internship_start=fromdate,internship_end=todate,fresher_experience=ele[12],previous_experience=ele[13],company_name=ele[14],register=ele[15],duration=ele[16],executive=usr,ex_duration=ele[17])
                            
        # file_up=request.FILES.get('up_ld_excel',None)
        # excel_contents = file_up.read_excel()

        
        try:
            file = request.FILES.get('up_ld_excel',None)
            wb = load_workbook(file)
            sheet = wb.active
            count = 0
            for row in sheet.iter_rows(values_only=True):

                print(row[16])
                print(row[18])
                print(row[9])
                print(row[8])
                if count == 0:
                    pass
                else:
                    if daily_leeds.objects.filter(name=row[2],email_id=row[1],ph_no=row[3]).exists():
                        if str(row[9]) == "no":
                            if str(row[8]) == "fresher":
                            
                                created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                            else:
                                fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)

                        else:
                            if str(row[8]) == "fresher":
                                fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                            else:
                                fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)
                    else:
                        

                        if int(row[16]) > 6:
                            if str(row[9]) == "no":
                                if str(row[8]) == "fresher":
                                
                                    created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                                else:
                                    fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)

                            else:
                                if str(row[8]) == "fresher":
                                    fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                                else:
                                    fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                    created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)
                        else:
                            
                            if int(row[18]) > 12:
                                if str(row[9]) == "no":
                                    if str(row[8]) == "fresher":
                                    
                                        created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)

                                else:
                                    if str(row[8]) == "fresher":
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None,executive=usr)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds_exists.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]),executive=usr)
                            else:

                                if str(row[9]) == "no":
                                    if str(row[8]) == "fresher":
                                    
                                        created = daily_leeds.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]))

                                else:
                                    if str(row[8]) == "fresher":
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],daily=dl,ex_duration=None)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = daily_leeds.objects.get_or_create(name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],daily=dl,ex_duration=int(row[18]))

                    
                    if All_leads.objects.filter(name=row[2],email_id=row[1],ph_no=row[3]).exists():
                        pass
                    else:
                        if int(row[16]) > 6:
                            pass
                        else:
                            
                            if int(row[18]) > 12:
                                pass
                            else:
                                if str(row[9]) == "no":
                                    if str(row[8]) == "fresher":
                                    
                                        created = All_leads.objects.get_or_create(date=date.today(),name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=None,company_name=None,duration=None, register=row[17],executive=usr,assign_status="no",ex_duration=None)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = All_leads.objects.get_or_create(date=date.today(),name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=None,internship_topic=None,internship_start=None,internship_end=None,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=None, register=row[17],executive=usr,assign_status="no",ex_duration=int(row[18]))

                                else:
                                    if str(row[8]) == "fresher":
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = All_leads.objects.get_or_create(date=date.today(),name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=None,company_name=None,duration=int(row[16]), register=row[17],executive=usr,assign_status="no",ex_duration=None)
                                    else:
                                        fromdate=datetime.strptime(str(row[11].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        todate=datetime.strptime(str(row[12].date()), '%Y-%m-%d').strftime('%Y-%m-%d')
                                        created = All_leads.objects.get_or_create(date=date.today(),name=row[2],email_id=row[1],ph_no=row[3],location=row[7],qualification=row[4],year_of_passout=int(row[5]),collegename=row[6],internship=row[9],internship_institute=row[13],internship_topic=row[10],internship_start=fromdate,internship_end=todate,fresher_experience=row[8],previous_experience=row[14],company_name=row[15],duration=int(row[16]), register=row[17],executive=usr,assign_status="no",ex_duration=int(row[18]))


                count+=1
        except:
            pass
    
        
       
        if dl.task == "Leads Collection":
            dls = daily_work.objects.filter(work=id)
            dl_cnt=0
            for l in dls:
                dl_cnt+=int(daily_leeds.objects.filter(daily=l.id).count())
            work.achived = dl_cnt
            work.save()

        
            if int(dl_cnt) > int(int(work.target)-10):
                if int(dl_cnt) == int(work.target) or int(dl_cnt) > int(work.target):
                    dl.status = "yes"
                    dl.target_count=dl_cnt
                    dl.status_date=date.today()
                    delays=dl.status_date - work.end_date
                    work.delay=int(delays.days)
                    work.status = "yes"
                    work.save()


                    leads=lead_delay()
                    leads.executive=usr
                    leads.date=date.today()
                    
                    leads.target=work.target
                    leads.status="yes"
                    leads.balance=int(work.target)-int(dl_cnt)
                    leads.achived=dl_cnt
                    leads.sub_target=int(work.target)-10
                    leads.save()
                else:
                    dl.status="about"
                    leads=lead_delay()
                    leads.executive=usr
                    leads.date=date.today()
                    dl.target_count=dl_cnt
                    leads.target=work.target
                    leads.status="about"
                    leads.balance=int(work.target)-int(dl_cnt)
                    leads.achived=dl_cnt
                    leads.sub_target=int(work.target)-10
                    leads.save()
                        

            elif int(dl_cnt) <= int(int(work.target)-10): 
                leads=lead_delay()
                leads.executive=usr
                leads.date=date.today()
                leads.target=work.target
                leads.status="no"
                
                leads.balance=int(work.target)-int(dl_cnt)
                leads.achived=dl_cnt
                leads.sub_target=int(work.target)-10
                leads.save()
                dl.status = "no"
                dl.target_count=dl_cnt   

            else:
                leads=lead_delay()
                leads.executive=usr
                leads.date=date.today()
                leads.target=work.target
                leads.status="about"
                
                leads.balance=int(work.target)-int(dl_cnt)
                leads.achived=dl_cnt
                leads.sub_target=int(work.target)-10
                leads.save()
                dl.status = "about"
                dl.target_count=dl_cnt
                
            dl.save()
        
        elif work.sub_task == "Off page":
            
            dls2 = daily_work.objects.filter(work=id)
            dlks=daily_work.objects.filter(work=work).count()
            if daily_off_sub.objects.filter(daily=dl).exists():
                dl_cnt2=0
                for l2 in dls2:
                    dl_cnt2+=int(daily_off_sub.objects.filter(daily=l2.id).count())
                dlk=int(dlks)+int(dl_cnt2)
            else:
                dlk=daily_work.objects.filter(work=work).count()

            work.achived = dlk
            work.save()

            if int(dlk) > int(int(work.target)-10):
                if ((int(dlk) == int(work.target)) or (int(dlk) > int(work.target))):
                    dl.status = "yes"
                    dl.status_date=date.today()
                    delays=dl.status_date - work.end_date
                    work.delay=int(delays.days)
                    work.status = "yes"
                    work.save()
                else:
                    dl.status = "about"
                    


            elif int(dlk) <= int(int(work.target)-10) :
                dl.status = "no"
                
            else:
                dl.status = 'about' 

            dl.save()
        else:
            dls3 = daily_work.objects.filter(work=id).count()
          

            work.achived = dls3
            work.save()

            if int(dls3) > int(int(work.target)-10):
                if int(dls3) == int(work.target) or int(dls3) > int(work.target):
                    dl.status = "yes"
                    dl.status_date=date.today()
                    delays=dl.status_date - work.end_date
               
                    work.delay=int(delays.days)
                    work.status = "yes"
                    work.save()
                else:
                   dl.status="about"  
                
            elif int(dls3) <= int(int(work.target)-10):
                dl.status = "no"
            else:
                dl.status = 'about'    
            dl.save()

            ###################################################

      
       
    
        return redirect("ex_daily_work_det",work.client_name_id)
    return redirect("ex_daily_work_det",work.client_name_id)

###################################################auto assign Section


def ex_delay(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    # dl_task=lead_delay.objects.filter(executive=usr)
    w_mails = dm_warning_mails.objects.filter(executive=usr)
    wk_assign = work_asign.objects.filter(exe_name=usr)
    w_id = [obj.work_id for obj in wk_assign]
    status_values = ["no"," "]
    work = Work.objects.filter(id__in=w_id,status__in=status_values).order_by('-start_date')
    print(work)
    for i in work:
        if i.target is not None:
            i.difference = i.target - 10
        else:
            i.difference = None

    today = date.today()
    context={
        "usr":usr,
        "work":work,
        'today':today,
        'w_mails':w_mails
    }
    return render(request, 'executive/ex_delay.html', context)


def ex_delay_det_day(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    if date:
        # dl_task=lead_delay.objects.filter(executive=usr,date=day_field)
        wk_assign = work_asign.objects.filter(exe_name=usr)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,status__in=status_values,start_date=day_field).order_by('-id')
        print(work)
        for i in work:
            if i.target is not None:
                i.difference = i.target - 10
            else:
                i.difference = None

        today = date.today()
        
    else:
        dl_task=lead_delay.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        "work":work,
        'today':today,

    }
    return render(request, 'executive/ex_delay.html', context)


def ex_delay_det_month(request):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year


    if date:
        # dl_task=lead_delay.objects.filter(Q(date__year=selected_year_number) & Q(date__month=selected_month_number),executive=usr)
        wk_assign = work_asign.objects.filter(exe_name=usr)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,status__in=status_values,start_date__year=selected_year_number,
                                   start_date__month=selected_month_number).order_by('-id')
        print(work)
        for i in work:
            if i.target is not None:
                i.difference = i.target - 10
            else:
                i.difference = None

        today = date.today()
       
    else:
        dl_task=lead_delay.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        "work":work,
        'today':today,
       
    }
    return render(request, 'executive/ex_delay.html', context)


def ex_weekly_rep_clint(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work_as=work_asign.objects.filter(exe_name=ids).values('client_name_id').distinct()
    work=Work.objects.all()
    last=work_asign.objects.filter(exe_name=ids).last()
    cl=client_information.objects.all()
    context={
        "usr":usr,
        "work_as":work_as,
        "work":work,
        "cl":cl,
        "last":last
    }
    return render(request, 'executive/ex_weekly_rep_clint.html',context)

def ex_weekly_rep_clint_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work_as=work_asign.objects.filter(exe_name=ids)
   
    work_tb=Work.objects.filter(client_name_id=id).order_by("-id")
    rep=progress_report.objects.filter(user=usr)
    context={
        "usr":usr,
        "work_as":work_as,
      
        "rep":rep,
        "work_tb":work_tb
    }
    return render(request, 'executive/ex_weekly_rep_det.html',context)

def sv_wk_rp(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    work=Work.objects.get(id=id)
    if request.method == 'POST':
        pro = progress_report()
        pro.task=work.task
        pro.audit_rprt=request.FILES.get('repr_fl',None)
        pro.graph=request.FILES.get('gr_fl',None)
        pro.start_date=request.POST.get('st_dt',None)
        pro.end_date=request.POST.get('ed_dt',None)
        pro.work=work
        pro.user=usr
        pro.cl_name=work.cl_name
        pro.save()
        return redirect("ex_weekly_rep_clint_det",work.client_name_id)
    return redirect("ex_weekly_rep_clint_det",work.client_name_id)

def ex_view_work_clint(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    work_as=work_asign.objects.filter(exe_name=ids)
    work=Work.objects.all()
    cl=client_information.objects.all()
    last=work_asign.objects.filter(exe_name=ids).last()

    curent_month= date.today().month
    curret_year= date.today().year

    context={
        "usr":usr,
        "work":work,
        "work_as":work_as,
        "cl":cl,
        "last":last,
    }
    return render(request, 'executive/ex_view_work_clint.html',context)

################## Auto Assign Section

now=datetime.now()
try:
    if now.hour == 24:
    
        curent_month= date.today().month
        curret_year= date.today().year
        
        target=Work.objects.filter(task="Leads Collection")
        dily=daily_work.objects.filter(task="Leads Collection")
    
        for i in target:
            if i.end_date < date.today():
                pass
            else:
                i.delay=0
                i.status="no"
                i.save()

        for k in dily:
            dls=Work.objects.get(id=k.work_id)
        
            if dls.end_date < date.today():
                pass
            else:
                k.status_date=None
                k.status="no"
                k.target_count=0
                k.save()
    else:
        pass
except:
    pass

def ex_view_clint_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    work=Work.objects.get(id=id)
    context={
        "usr":usr, 
        "client":work
    }
    return render(request, 'executive/ex_view_clint_det.html',context)

def ex_warnings_dash(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr
    }
    return render(request, 'executive/ex_warnings_dash.html',context) 


def ex_warning(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)

    warn=Warning.objects.filter(executive=ids,type="Warning")
    context={
        "usr":usr,
        "warn":warn
    }
    return render(request, 'executive/ex_warning.html',context)

def add_warning(request, id):
   

    if request.method == 'POST':
        warn = Warning.objects.get(id=id)
        warn.reply=request.POST.get('workdone',None)
        warn.save()
        return redirect("ex_warning")
    return redirect("ex_warning")
    
def ex_suggestions(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    warn=Warning.objects.filter(executive=ids,type="Suggestion")
    context={
        "usr":usr,
        "warn":warn
    }
    return render(request, 'executive/ex_suggestions.html',context)

def add_suggestion(request, id):
   

    if request.method == 'POST':
        warn = Warning.objects.get(id=id)
        warn.reply=request.POST.get('workdone',None)
        warn.save()
        return redirect("ex_warning")
    return redirect("ex_warning")


def corrections(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    cor=correction.objects.filter(executive=ids)
    dl_work=daily_work.objects.filter(user=ids)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dl_leeds=daily_leeds.objects.all()
    print(dl_leeds)
    context={
        "usr":usr,
        "warn":cor,
        "dl_work":dl_work,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,

    }
    return render(request, 'executive/correction.html',context)

def add_corrections(request, id):
   

    if request.method == 'POST':
        warn = correction.objects.get(id=id)
        warn.reply=request.POST.get('workdone',None)
        warn.save()
        return redirect("corrections")
    return redirect("corrections")

def get_corrections(request):
    ele = request.GET.get('ele')
    warn = correction.objects.get(id=ele)
    warns =warn.description
    rep =warn.reply
 
    return JsonResponse({"status":" not","warns":warns,"rep":rep})

    
def get_warns(request):
    ele = request.GET.get('ele')
    warn = Warning.objects.get(id=ele)
    warns =warn.description
    rep =warn.reply
 
    return JsonResponse({"status":" not","warns":warns,"rep":rep})

    
def get_requ(request):
    ele = request.GET.get('ele')
    warn = addi_client_info.objects.get(id=ele)
    warns =warn.discription
    rep =warn.file
    nm =warn.labels
    target =warn.target
   
    vk=str(rep)
    
    return JsonResponse({"status":" not","warns":warns,"rep":vk,"nm":nm,"target":target})


def logout(request):
    if 'userid' in request.session:  
        request.session.flush()
        return redirect('/')
    else:
        return redirect('/') 


def ex_change_pass(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')

    dev = user_registration.objects.filter(id=devid)

    if request.method == 'POST':
        abc = user_registration.objects.get(id=devid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'executive/ex_ch_pass.html', {'dev': dev,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'executive/ex_ch_pass.html', {'dev': dev,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'executive/ex_ch_pass.html', {'dev': dev,"usr":usr})
    return render(request, 'executive/ex_ch_pass.html', {'dev': dev,"usr":usr})

def ex_accountset(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    return render(request, 'executive/ex_accountset.html', {'dev': dev,"usr":usr})

def ex_imagechange(request, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    if request.method == 'POST':
        abc = user_registration.objects.get(id=id)
        abc.photo = request.FILES['filename']
        
        abc.save()
        return redirect('ex_accountset')
    return render(request, 'executive/ex_accountset.html',{'dev': dev,"usr":usr})


def ex_schedule_dash(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    
    }
    return render(request, 'executive/ex_schedule_dash.html',context)

def ex_calander(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    
    }
    return render(request, 'executive/ex_calander.html',context)


def ex_all_events(request):
    all_events = Events.objects.all()
    out=[]
    for event in all_events:
        out.append({
            "title":event.name,
            "id":event.id,
            "start":event.start.strftime("%m/%d/%Y, %H:%M:%S"), 
        })
    return JsonResponse(out, safe=False)
 
 
def ex_add_event(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method == 'POST':
        start = request.POST.get('start', None)
        end = request.POST.get("end", None)
        title = request.POST.get('title', None)
        img = request.FILES.get('file', None)

        b = Events()
        b.name=title
        b.start=start 
        b.end=end 
        b.img=img
        b.executive=usr
        b.status="pending"
        b.save()
        data = {}
        return JsonResponse(data)
 
def ex_update(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.start = start
    event.end = end
    event.name = title
    event.save()
    data = {}
    return JsonResponse(data)
 
def ex_remove(request):
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)

def ex_shedule_work(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    events = Events.objects.filter(executive=usr)
    dl_sub = addi_events.objects.filter(executive=usr)

    
    context={
        "usr":usr,
        'events': events,
        "dl_sub":dl_sub

    }
    return render(request, 'executive/ex_schedule_work.html',context)


def ex_edit_post_status(request,id):
    if request.method == 'POST':
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        b=Events.objects.get(id=id)
        b.status=request.POST['status']
        b.st_file=request.FILES.get('cmpl_file',None)
        b.fb = request.POST.get('fb',None)
        b.fb_dt = request.POST.get('fb_txt',None)
        b.fb_file = request.FILES.get('fb_file',None)
        b.tw = request.POST.get('tw',None)
        b.tw_dt = request.POST.get('tw_txt',None)
        b.tw_file = request.FILES.get('tw_file',None)
        b.pin = request.POST.get('pin',None)
        b.pin_dt = request.POST.get('pin_txt',None)
        b.pin_file = request.FILES.get('pin_file',None)
        b.link = request.POST.get('link',None)
        b.link_dt = request.POST.get('link_txt',None)
        b.link_file = request.FILES.get('link_file',None)
        b.insta = request.POST.get('insta',None)
        b.insta_dt = request.POST.get('insta_txt',None)
        b.insta_file = request.FILES.get('insta_file',None)
        b.tumb = request.POST.get('tumb',None)
        b.tumb_dt = request.POST.get('tumb_txt',None)
        b.tumb_file = request.FILES.get('tumb_file',None)
        b.diry = request.POST.get('diry',None)
        b.diry_dt = request.POST.get('diry_txt',None)
        b.diry_file = request.FILES.get('diry_file',None)
        b.yt = request.POST.get('yt',None)
        b.yt_dt = request.POST.get('yt_txt',None)
        b.yt_file = request.FILES.get('yt_file',None)
        b.qra = request.POST.get('qra',None)
        b.qra_dt = request.POST.get('qra_txt',None)
        b.qra_file = request.FILES.get('qra_file',None)
        b.sbms = request.POST.get('sbms',None)
        b.sbms_dt = request.POST.get('sbms_txt',None)
        b.sbms_file = request.FILES.get('sbms_file',None)
        
        b.save()

        
        label_req =request.POST.getlist('sub_lb[]')
        dt =request.POST.getlist('dates[]') 
        files_req =request.FILES.getlist('sub_file[]') 
      

        
        if len(files_req)==len(label_req)==len(dt):
            mapped2 = zip(label_req,dt,files_req)
            mapped2=list(mapped2)
         
            for ele in mapped2:
                created = addi_events.objects.get_or_create(label=ele[0],date=ele[1],file=ele[2],executive=usr,events=b)
                
      
        b.save()
        return redirect('ex_shedule_work')

    return redirect('ex_shedule_work')
    
def ex_save_shedule(request):
    idr=request.session['userid']
    usr_lg = user_registration.objects.get(id=idr)

    if request.method == 'POST':

        ids=request.session['smo_userid']
        usr = smo_registration.objects.get(id=ids)
        b=smo_post()
        b.description = request.POST['description']
        b.status="pending"
        dct_file = dict(request.FILES)
        lst_screenshot = dct_file['filed']
        lst_file = []
        for ins_screenshot in lst_screenshot:
            str_img_path = ""
            if ins_screenshot:
                img_emp = ins_screenshot
                fs = FileSystemStorage(location=settings.MEDIA_ROOT,base_url=settings.MEDIA_URL)
                str_img = fs.save(''.join(filter(str.isalnum, str(img_emp))), img_emp)
                str_img_path = fs.url(''.join(filter(str.isalnum, str_img)))
                lst_file.append('/media/'+''.join(filter(str.isalnum, str(img_emp))))
                b.json_testerscreenshot = lst_file

        
        b.json_testerscreenshot=b.json_testerscreenshot
        b.smo=usr
        b.executive=usr_lg

 
        b.fb = request.POST.get('fb',None)
        b.fb_dt = request.POST.get('fb_txt',None)
        b.fb_file = request.FILES.get('fb_file',None)
        b.tw = request.POST.get('tw',None)
        b.tw_dt = request.POST.get('tw_txt',None)
        b.tw_file = request.FILES.get('tw_file',None)
        b.pin = request.POST.get('pin',None)
        b.pin_dt = request.POST.get('pin_txt',None)
        b.pin_file = request.FILES.get('pin_file',None)
        b.link = request.POST.get('link',None)
        b.link_dt = request.POST.get('link_txt',None)
        b.link_file = request.FILES.get('link_file',None)
        b.insta = request.POST.get('insta',None)
        b.insta_dt = request.POST.get('insta_txt',None)
        b.insta_file = request.FILES.get('insta_file',None)
        b.tumb = request.POST.get('tumb',None)
        b.tumb_dt = request.POST.get('tumb_txt',None)
        b.tumb_file = request.FILES.get('tumb_file',None)
        b.diry = request.POST.get('diry',None)
        b.diry_dt = request.POST.get('diry_txt',None)
        b.diry_file = request.FILES.get('diry_file',None)
        b.yt = request.POST.get('yt',None)
        b.yt_dt = request.POST.get('yt_txt',None)
        b.yt_file = request.FILES.get('yt_file',None)
        b.qra = request.POST.get('qra',None)
        b.qra_dt = request.POST.get('qra_txt',None)
        b.qra_file = request.FILES.get('qra_file',None)
        b.sbms = request.POST.get('sbms',None)
        b.sbms_dt = request.POST.get('sbms_txt',None)
        b.sbms_file = request.FILES.get('sbms_file',None)
        
        b.save()

        
        label_req =request.POST.getlist('sub_lb[]')
        dt =request.POST.getlist('dates[]') 
        files_req =request.FILES.getlist('sub_file[]') 
      

        
        if len(files_req)==len(label_req)==len(dt):
            mapped2 = zip(label_req,dt,files_req)
            mapped2=list(mapped2)
         
            for ele in mapped2:
                created = addi_smo_post.objects.get_or_create(label=ele[0],date=ele[1],file=ele[2],executive=usr_lg,smo=usr,post=b)
        return redirect('ex_calander')
    return redirect('ex_calander')


def leave_home(request):
    if 'userid' in request.session:
        
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        
     
        return render(request, 'executive/leave_home.html',{'usr':usr})
    else:
        return redirect('/')

def leave_aply(request):
    if 'userid' in request.session:
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        return render(request, 'executive/leave_apply.html',{'usr':usr})
    else:
        return redirect('/')

def ex_leave_form(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        des1 = user_registration.objects.get(id=ids)
        if request.method == "POST":
            leaves = leave()
            leaves.from_date = request.POST['from']
            leaves.to_date = request.POST['to']
            leaves.leave_status = request.POST['haful']
            leaves.reason = request.POST['reason']
            leaves.user_id = ids
            leaves.status = "submitted"
            leaves.designation_id = des1.id
            leaves.leaveapprovedstatus=0
            
            start = datetime.strptime(leaves.from_date, '%Y-%m-%d').date() 
            end = datetime.strptime(leaves.to_date, '%Y-%m-%d').date()

            diff = (end  - start).days
    
            leaves.save()
        return  redirect('ex_all_leave')  
    else:
        return redirect('/')

def ex_all_leave(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        var = leave.objects.filter(user_id=ids)
        return render(request, 'executive/ex_all_leave.html',{'var': var,'usr':usr})
    else:
        return redirect('/')

def duplicate_leads(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    today=date.today()
    usr = user_registration.objects.get(id=ids)
    count=daily_leeds_exists.objects.filter(date=today,executive=usr).count()

    duplicate= daily_leeds_exists.objects.filter(executive=usr).order_by('date')
    
    context={
        "usr":usr,
        'duplicate':duplicate,
        'count':count,
    }
    return render(request,'executive/ex_duplicate_leads.html',context)


def ex_flt_day_duplicate(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    day_field=request.POST.get('day')

    duplicate= daily_leeds_exists.objects.filter(date=day_field,executive=usr).order_by('-date')
    

    
    context={
        "usr":usr,
        'duplicate':duplicate,
    }
    return render(request, 'executive/ex_duplicate_leads.html',context)


def ex_flt_month_duplicate(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year


    duplicate=daily_leeds_exists.objects.filter(Q(date__year=selected_year_number) & Q(date__month=selected_month_number),executive=usr).order_by('-date')

    
    
    
    context={
        "usr":usr,
        'duplicate':duplicate
       

    }
    return render(request, 'executive/ex_duplicate_leads.html',context)

#---------------------------------marketing head section

    

def he_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'head/he_profile.html',{"usr":usr})

def he_project(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request,'head/he_project.html',{"usr":usr})

def he_view_works(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.all()
    return render(request,'head/he_view_works.html',{'client':client,"usr":usr,})

def he_work_asign(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    exe=user_registration.objects.filter(department='Digital Marketing Executive')
    return render(request,'head/he_work_asign.html',{'client':client,'exe':exe,"usr":usr,})

def he_view_work_asign_client(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.all()
    return render(request,'head/he_view_work_asign_client.html',{'client':client,"usr":usr,})

def he_view_work_asign_exe(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=id)
    work=work_asign.objects.filter(client_name=client.id)
    return render(request,'head/he_view_work_asign_exe.html',{"usr":usr,"w":work})


def he_daily_task(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today=date.today()
    work=daily_work.objects.filter(date=today)
    master=Work.objects.all()
    sub_work=daily_work_sub.objects.all()
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()

    dl_leeds=daily_leeds.objects.all()
    return render(request,'head/he_daily_task.html',{'work':work,"usr":usr,"sub":sub_work,"dl_sub":dl_sub,"dl_off":dl_off,"dl_leeds":dl_leeds,'master':master})


def he_workprogress_executive(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    prgs=progress_report.objects.all()
    return render(request,'head/he_workprogress_executive.html',{'prgs':prgs,"usr":usr,})

def he_progress_report(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work=progress_report.objects.get(id=pk)
    try:
        prv_work=progress_report.objects.filter(work_id=work.id).order_by('-end_date')[0]
    except:
        prv_work=None
    return render(request,'head/he_progress_report.html',{'work':work,"usr":usr,"prv_work":prv_work})


def he_feedback(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request,'head/he_feedback.html',{'exe':exe,"usr":usr,})


def he_feedbacke1(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.get(id=pk)
    wrng=Warning.objects.filter(executive_id=exe.id)
    return render(request,'head/he_feedback1.html',{'exe':exe,'wrng':wrng,"usr":usr,})

    
def he_feedback_submit(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method=='POST':
        des=request.POST['des']
        typ=request.POST['option']
        warning=Warning(executive_id=pk,description=des,type=typ,)
        warning.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def he_work_add(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method == 'POST':
        task = request.POST.get('task')
        des = request.POST.get('des')
        sdate=request.POST.get('sdate')
        edate=request.POST.get('edate')
        file=request.FILES.get('file')
        sub_tsk = request.POST.get('sub_tsk')
        target=request.POST.get('trgt')
        client=client_information.objects.get(id=id)
        json_data = request.POST.get('array', '')
        array = json.loads(json_data)
        w=Work(task=task,description=des,start_date=sdate,end_date=edate,file_attached=file,cl_name=client.bs_name,client_name=client)
        w.save()

        if w.task=="SEO":
            w.file_2=client.seo_file
            w.save()
            if sub_tsk=="On page":
                w.sub_task="On page"
                w.sub_des=client.on_pg_txt
                w.sub_file=client.on_pg_file
                w.target=request.POST.get('ontrgt')
                w.save()
            if sub_tsk=="Off page":
                w.sub_task="Off page"
                w.sub_des=client.off_pg_txt
                w.sub_file=client.off_pg_file
                w.target=request.POST.get('offtrgt')
                w.save()

        if w.task=="SMM":
            w.file_2=client.smm_file
            w.target=target
            w.save()
        if w.task=="SEM/PPC":
            w.file_2=client.sem_file
            w.target=target
            w.save()
        if w.task=="Email Marketing":
            w.file_2=client.em_file
            w.target=target
            w.save()   
        if w.task=="Content Marketing":
            w.file_2=client.cm_file
            w.target=target
            w.save() 
        if w.task=="Affiliate Marketing":
            w.file_2=client.am_file
            w.target=target
            w.save()   
        if w.task=="Mobile marketing":
            w.file_2=client.mm_file
            w.target=target
            w.save()  
        if w.task=="Video Marketing":
            w.file_2=client.vm_file
            w.target=target
            w.save() 
        if w.task=="SMO":
            w.file_2=client.smo_file
            w.target=target
            w.save() 
        if w.task=="Leads Collection":
            w.file_2=client.lc_file
            w.target=target
            w.save()            
        w=Work.objects.latest('id')
        for i in array:
            b=user_registration.objects.get(department="Digital Marketing Executive",fullname=i)
            c=work_asign(work_id=w.id,exe_name_id=b.id,client_name_id=client.id)
            c.save()
        return HttpResponse({"message": "success"})


def he_add_correction_daily(request,id):
    if request.method=='POST':
        cor=request.POST.get('cor')
        daily=daily_work.objects.get(id=id)
        print(daily.user_id)
        c=correction(description=cor,daily_id=daily.id,executive_id=daily.user_id,event=None)
        c.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def he_change_pass(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)

    if request.method == 'POST':
        abc = user_registration.objects.get(id=devid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'head/he_ch_pass.html', {'dev': dev,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'head/he_ch_pass.html', {'dev': dev,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'head/he_ch_pass.html', {'dev': dev,"usr":usr})
    return render(request, 'head/he_ch_pass.html', {'dev': dev,"usr":usr})


def he_accountset(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    return render(request, 'head/he_accountset.html', {'dev': dev,"usr":usr})

def he_imagechange(request, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    if request.method == 'POST':
        abc = user_registration.objects.get(id=id)
        abc.photo = request.FILES['filename']
        abc.save()
        return redirect('he_accountset')
    return render(request, 'head/he_accountset.html',{'dev': dev,"usr":usr})


def he_flt_progress(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
  
    pr_work=progress_report.objects.filter(start_date__gte=st_dt,start_date__lte=en_dt)
    context={
        "usr":usr,
        "prgs":pr_work

    }
    return render(request, 'head/he_workprogress_executive.html',context)

def he_view_post(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today = date.today()
    event=Events.objects.filter(executive=id, start__date=today)
    hr_list=["09:00:00","10:00:00","11:00:00","12:00:00","13:00:00","14:00:00","15:00:00","16:00:00"]
    start_values = [i.start.strftime("%H:%M:%S") for i in event]
    unschedule = [datetime.strptime(elem, '%H:%M:%S').strftime('%I:%M %p') for elem in hr_list if elem not in start_values]
    return render(request,'head/he_view_post.html',{"usr":usr,'evnt':event,'unsh':unschedule})

def he_add_correction(request,id):
    if request.method=='POST':
        cor=request.POST.get('cor')
        event=Events.objects.get(id=id)
        c=correction(description=cor,executive_id=event.executive.id,event_id=event.id,daily=None)
        c.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

        
def he_add_status(request,id):
    if request.method=='POST':
        status=request.POST.get('status')   
        daily=daily_work.objects.get(id=id)
        daily.status=status
        daily.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def he_add_event_status(request,id):
    if request.method=='POST':
        status=request.POST.get('status')   
        event=Events.objects.get(id=id)
        event.status=status
        event.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
    

def he_smo_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request, 'head/he_smo_exe.html',{'exe':exe,"usr":usr})

def he_cor_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request, 'head/he_cor_exe.html',{'exe':exe,"usr":usr})

def he_cor_exe_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    cor=correction.objects.filter(executive=id)
    

    context={
        "usr":usr,
        "warn":cor,
        

    }
    return render(request, 'head/he_cor_exe_det.html',context)

def he_leave_home(request):
    if 'userid' in request.session:
        
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        
     
        return render(request, 'head/he_leave_home.html',{'usr':usr})
    else:
        return redirect('/')

def he_leave_aply(request):
    if 'userid' in request.session:
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        return render(request, 'head/he_leave_apply.html',{'usr':usr})
    else:
        return redirect('/')

def he_leave_form(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        des1 = user_registration.objects.get(id=ids)
        if request.method == "POST":
            leaves = leave()
            leaves.from_date = request.POST['from']
            leaves.to_date = request.POST['to']
            leaves.leave_status = request.POST['haful']
            leaves.reason = request.POST['reason']
            leaves.user_id = ids
            leaves.status = "submitted"
            leaves.designation_id = des1.id
            leaves.leaveapprovedstatus=0
            
            start = datetime.strptime(leaves.from_date, '%Y-%m-%d').date() 
            end = datetime.strptime(leaves.to_date, '%Y-%m-%d').date()

            diff = (end  - start).days
    
            leaves.save()
        return  redirect('he_all_leave')  
    else:
        return redirect('/')

def he_all_leave(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        var = leave.objects.filter(user_id=ids)
        return render(request, 'head/he_all_leave.html',{'var': var,'usr':usr})
    else:
        return redirect('/')


#-------------------------------------------------------------------------------Smo Submission
def smo_base(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
   
    

    context={
        "usr":usr,
      
    }
    return render(request, 'smo/publishing/smo_base.html',context)

def smo_login(request,id):
    work=Work.objects.get(id=id)
    return render(request, 'smo/index/smo_login.html', {'id':work.client_name_id})

def smo_signup(request,id):
    return render(request, 'smo/index/smo_signup.html', {'id':id})



def smo_reg(request,id):

 
    a = smo_registration()
    
    
    
    client=client_information.objects.get(id=id)
    if request.method == 'POST':
        if  smo_registration.objects.filter(email=request.POST['email']).exists():
            
            msg_error = "Mail id already exist"
            return render(request, 'smo/index/smo_signup.html',{'msg_error': msg_error})
        else:
            if request.POST['password'] == request.POST['re_password']:
                a.fullname = request.POST['fname']
                a.email = request.POST['email']
                a.password = request.POST['password']
                a.photo = request.FILES['photo']
                a.client=client
                a.save()
                return redirect('smo_login',id)
            else:
                msg_error = "Mail id already exist"
                return render(request, 'smo/index/smo_signup.html',{'msg_error': msg_error, "id": id})

def smo_signin(request,id):  
    if request.method == 'POST':
        email  = request.POST['email']
        password = request.POST['password']
        
        if smo_registration.objects.filter(email=request.POST['email'], password=request.POST['password']).exists():
            smo_ex = smo_registration.objects.get(email=request.POST['email'],password=request.POST['password'])

            #---------------------- executive session id
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
            request.session['userid'] = usr.id
            #---------------------- smo submission login session id
            request.session['smo_userid'] = smo_ex.id
            
            return redirect('smo_dash')
    return redirect('smo_login',id) 

def smo_dash(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
    context={
            "usr":usr,
        }
    return render(request, 'smo/publishing/smo_dashboard.html',context)

def smo_cnt_chnl(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
    context={
            "usr":usr,
        }
    return render(request, 'smo/publishing/connect_channel.html',context) 


def published_post(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
    context={
            "usr":usr,
        }
    return render(request, 'smo/publishing/published_post.html',context)


def create_post(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids) 
    post = smo_post.objects.filter(smo=usr)
    addi_post = addi_smo_post.objects.all()
    
    dt=date.today()
    context={
            "usr":usr,
            "post":post,
            "dt":dt,
            "addi_post":addi_post
        }
    return render(request, 'smo/publishing/create_post.html',context)


def edit_post_drft(request,id):
    if request.method == 'POST':
        ids=request.session['smo_userid']
        usr = smo_registration.objects.get(id=ids)
        b=smo_post.objects.get(id=id)
        b.description = request.POST['description']
        b.status=request.POST['status']
        b.st_file=request.FILES.get('cmpl_file',None)
        if request.FILES.get('filed',None) == None:
            pass
        else:
            dct_file = dict(request.FILES)
            try:
                lst_screenshot = dct_file['filed']
                lst_file = []
                for ins_screenshot in lst_screenshot:
                    str_img_path = ""
                    if ins_screenshot:
                        img_emp = ins_screenshot
                        fs = FileSystemStorage(location=settings.MEDIA_ROOT,base_url=settings.MEDIA_URL)
                        str_img = fs.save(''.join(filter(str.isalnum, str(img_emp))), img_emp)
                        str_img_path = fs.url(''.join(filter(str.isalnum, str_img)))
                        lst_file.append('/media/'+''.join(filter(str.isalnum, str(img_emp))))
                        b.json_testerscreenshot = lst_file
            except:
                b.json_testerscreenshot=b.json_testerscreenshot
        b.smo=usr
     
        b.save()
        return redirect('create_post')

    return redirect('create_post')


def save_post_drft(request):
    idr=request.session['userid']
    usr_lg = user_registration.objects.get(id=idr)

    if request.method == 'POST':

        ids=request.session['smo_userid']
        usr = smo_registration.objects.get(id=ids)
        b=smo_post()
        b.description = request.POST['description']
        b.status="draft"
        dct_file = dict(request.FILES)
        lst_screenshot = dct_file['filed']
        lst_file = []
        for ins_screenshot in lst_screenshot:
            str_img_path = ""
            if ins_screenshot:
                img_emp = ins_screenshot
                fs = FileSystemStorage(location=settings.MEDIA_ROOT,base_url=settings.MEDIA_URL)
                str_img = fs.save(''.join(filter(str.isalnum, str(img_emp))), img_emp)
                str_img_path = fs.url(''.join(filter(str.isalnum, str_img)))
                lst_file.append('/media/'+''.join(filter(str.isalnum, str(img_emp))))
                b.json_testerscreenshot = lst_file

        
        b.json_testerscreenshot=b.json_testerscreenshot
        b.smo=usr
        b.executive=usr_lg


        b.fb = request.POST.get('fb',None)
        b.fb_dt = request.POST.get('fb_txt',None)
        b.fb_file = request.FILES.get('fb_file',None)
        b.tw = request.POST.get('tw',None)
        b.tw_dt = request.POST.get('tw_txt',None)
        b.tw_file = request.FILES.get('tw_file',None)
        b.pin = request.POST.get('pin',None)
        b.pin_dt = request.POST.get('pin_txt',None)
        b.pin_file = request.FILES.get('pin_file',None)
        b.link = request.POST.get('link',None)
        b.link_dt = request.POST.get('link_txt',None)
        b.link_file = request.FILES.get('link_file',None)
        b.insta = request.POST.get('insta',None)
        b.insta_dt = request.POST.get('insta_txt',None)
        b.insta_file = request.FILES.get('insta_file',None)
        b.tumb = request.POST.get('tumb',None)
        b.tumb_dt = request.POST.get('tumb_txt',None)
        b.tumb_file = request.FILES.get('tumb_file',None)
        b.diry = request.POST.get('diry',None)
        b.diry_dt = request.POST.get('diry_txt',None)
        b.diry_file = request.FILES.get('diry_file',None)
        b.yt = request.POST.get('yt',None)
        b.yt_dt = request.POST.get('yt_txt',None)
        b.yt_file = request.FILES.get('yt_file',None)
        b.qra = request.POST.get('qra',None)
        b.qra_dt = request.POST.get('qra_txt',None)
        b.qra_file = request.FILES.get('qra_file',None)
        b.sbms = request.POST.get('sbms',None)
        b.sbms_dt = request.POST.get('sbms_txt',None)
        b.sbms_file = request.FILES.get('sbms_file',None)
        
        b.save()

        
        label_req =request.POST.getlist('sub_lb[]')
        dt =request.POST.getlist('dates[]') 
        files_req =request.FILES.getlist('sub_file[]') 
      

        
        if len(files_req)==len(label_req)==len(dt):
            mapped2 = zip(label_req,dt,files_req)
            mapped2=list(mapped2)
         
            for ele in mapped2:
                created = addi_smo_post.objects.get_or_create(label=ele[0],date=ele[1],file=ele[2],executive=usr_lg,smo=usr,post=b)
        return redirect('create_post')
    return redirect('create_post')


def content(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids) 
    post = smo_post.objects.filter(smo=usr)
    addi_post=addi_smo_post.objects.all()

    context={
            "usr":usr,
            "post":post,
            "addi_post":addi_post
        }
    return render(request, 'smo/publishing/content.html',context)


def logout_smo(request):
    if 'smo_userid' in request.session:  
        request.session.flush()
        return redirect('/')
    else:
        return redirect('/') 

def smo_change_pass(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
    if request.session.has_key('smo_userid'):
        devid = request.session['smo_userid']
    else:
        return redirect('/')
    dev = smo_registration.objects.filter(id=devid)

    if request.method == 'POST':
        abc = smo_registration.objects.get(id=devid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'smo/index/smo_change_password.html', {'dev': dev,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'smo/index/smo_change_password.html', {'dev': dev,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'smo/index/smo_change_password.html', {'dev': dev,"usr":usr})
    return render(request, 'smo/index/smo_change_password.html', {'dev': dev,"usr":usr})


def sm_calendar(request):
    ids=request.session['smo_userid']
    usr = smo_registration.objects.get(id=ids)
    all_events = Events.objects.all()
    context = {
        "events":all_events,
        "usr":usr,
    }
    return render(request, 'smo/publishing/calendar.html',context)


   
def all_events(request):
    all_events = Events.objects.all()
    out=[]
    for event in all_events:
        out.append({
            "title":event.name,
            "id":event.id,
            "start":event.start.strftime("%m/%d/%Y, %H:%M:%S"), 
        })
    return JsonResponse(out, safe=False) 
 
 
def add_event(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method == 'POST':
        start = request.POST.get('start', None)
        title = request.POST.get('title', None)
        img = request.FILES.get('file', None)
        date_obj = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
        one_hour = timedelta(hours=1)
        new_date_obj = date_obj + one_hour
        end = new_date_obj.strftime("%Y-%m-%d %H:%M:%S.%f")
        event = Events(name=title, start=start,end=end, img=img,executive=usr, status="draft") 
        event.save()
        data = {}
        return JsonResponse(data)
 
def update(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.start = start
    event.end = end
    event.name = title
    event.save()
    data = {}
    return JsonResponse(data)
 
def remove(request):
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)



# def Analysis(request):
#     if 'uid' in request.session:
#         if request.session.has_key('userid'):
#             uid = request.session['userid']
           
#         else:
#             return render(request,'Analsis_HomePage.html')
#         if request.method == "POST":
#             user_reg=user_registration.objects.get(id=userid)
            
#             file_up=UploadFiles.objects.get(id=pk)
#             excel_contents = file_up.read_excel()
#             return redirect('ex_dashboard')
   
#         return redirect('ex_dashboard')

#     else:
#         return redirect('ex_dashboard')

# from django.shortcuts import render
# from openpyxl import load_workbook

# def upload_excel(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data['file']
#             wb = load_workbook(file)
#             sheet = wb.active

#             # Process the data from the sheet
#             for row in sheet.iter_rows(values_only=True):
#                 # Here, you can perform operations with each row's data
#                 # For example, you can save the data to your SQL table using Django ORM

#             # Optionally, you can render a success message
#             return redirect('ex_dashboard')
#         else:
#             return redirect('ex_dashboard')
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'upload.html', {'form': form})



#....................................................<muhammed anas ><datamanager >...............................................................#

import random
from django.db.models import Count
from django.shortcuts import render, get_object_or_404
from .models import All_leads , Lead_assign


def dm_dashboard(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'dm_manager/dm_dashboard.html',  {'usr': usr})

def view_leads(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads_by_date = All_leads.objects.values('date').annotate(total_leads=Count('id'))
    for lead in leads_by_date:
        lead['target'] = random.randint(70, 120)
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date
    }
    return render(request, 'dm_manager/view_leads.html', context)




def view_all_leads(request, date=None):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if date:
        leads = All_leads.objects.filter(date=date)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads
    }
    return render(request, 'dm_manager/view_all_leads.html', context)




def assign(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    # leads_by_date = All_leads.objects.values('date').annotate(total_leads=Count('id'))
    status = ["no"," "]
    leads_by_date = All_leads.objects.values('date').annotate(total_leads=Count('id'), not_assigned=Count('id', filter=Q(assign_status__in=status)))

    
    context = {
         "usr":usr,
        'leads_by_date': leads_by_date
    }
    return render(request, 'dm_manager/assign.html', context)

from .models import user_registration

def assign_all(request, date=None):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if date:
        leads = All_leads.objects.filter(date=date)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided

    telecallers = user_registration.objects.filter(department='Telecaller')
    context = {
        "usr":usr,
        'leads': leads,
        'telecallers': telecallers
    }
    return render(request, 'dm_manager/assign_all.html', context)

from django.shortcuts import redirect

from datetime import date  # Add this import statement

def save_assignment(request):
    ids = request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method == 'POST':
        selected_leads = request.POST.getlist('lead_checkbox')
        selected_telecaller_name = request.POST.get('telecaller')

        if selected_telecaller_name and selected_leads:
            telecaller = user_registration.objects.get(fullname=selected_telecaller_name)
            already_assigned_leads = []
            for lead_id in selected_leads:
                lead = All_leads.objects.get(id=lead_id)
                if lead.assign_status == "yes":
                    already_assigned_leads.append(lead)
                else:
                    assignment = Lead_assign(telecaller=selected_telecaller_name, checkbox=lead)
                    assignment.save()
                    lead.assign_status = "yes"
                    lead.telecaller_id = telecaller.id
                    lead.data_manager_id = ids
                    lead.assign_dt = date.today()  # Set the assign_dt to the current date
                    lead.save()

            context = {
                "usr": usr,
                'already_assigned_leads': already_assigned_leads,
            }
            return render(request, 'dm_manager/assign_all.html', context)
        else:
            telecallers = user_registration.objects.filter(department='Telecaller')
            leads = All_leads.objects.none()
            context = {
                'leads': leads,
                'telecallers': telecallers,
                'error_message': "Please select a telecaller and at least one lead."
            }
            return render(request, 'dm_manager/assign_all.html', context)



from django.shortcuts import render
from .models import All_leads

def leads_view(request):
    leads = All_leads.objects.all()
    return render(request, 'view_.html', {'leads': leads})


def assign_count(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    telecallers = Lead_assign.objects.values('telecaller').annotate(assign_count=Count('telecaller'))
    context = {
         "usr":usr,
        'telecallers': telecallers
    }
    return render(request, 'dm_manager/assign_count.html', context)


def count_all(request, telecaller):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assigned_leads = Lead_assign.objects.filter(telecaller=telecaller)
    context = {
         "usr":usr,
        'assigned_leads': assigned_leads
    }
    return render(request, 'dm_manager/count_all.html', context)


from django.http import JsonResponse

def delete_lead(request, lead_id):
    if request.method == 'DELETE':
        try:
            lead = All_leads.objects.get(id=lead_id)
            lead.delete()
            return JsonResponse({'message': 'Lead deleted successfully'})
        except All_leads.DoesNotExist:
            return JsonResponse({'error': 'Lead not found'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)

def dm_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'dm_manager/dm_profile.html',context )

def all_counts(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assigned_lead_count = Lead_assign.objects.count()  # Count of assigned leads
    pending_lead_count = All_leads.objects.filter(telecaller_id=None).count()  # Count of pending leads
    context = {
        'assigned_lead_count': assigned_lead_count,
        "usr":usr,
        'pending_lead_count': pending_lead_count
    }
    return render(request, 'dm_manager/all_counts.html', context)

from django.shortcuts import render

def pending_count(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    pending_leads = All_leads.objects.filter(telecaller_id=None)  # Fetch leads not assigned to any telecaller
    context = {
        'pending_leads': pending_leads,
        "usr":usr,

    }
    return render(request, 'dm_manager/pending_count.html', context)

def dm_accountset(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.filter(id=callerid)
    return render(request, 'dm_manager/dm_accountset.html', {'caller': caller,"usr":usr})

def dm_change_pass(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')

    caller = user_registration.objects.filter(id=callerid)

    if request.method == 'POST':
        abc = user_registration.objects.get(id=callerid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'dm_manager/dm_pass.html', {'caller': caller,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'dm_manager/dm_pass.html', {'caller': caller,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'dm_manager/dm_pass.html', {'caller': caller,"usr":usr})
    return render(request, 'dm_manager/dm_pass.html', {'caller': caller,"usr":usr})

def dm_imagechange(request, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    if request.method == 'POST':
        abc = user_registration.objects.get(id=id)
        abc.photo = request.FILES['filename']
        
        abc.save()
        return redirect('dm_accountset')
    return render(request, 'dm_manager/dm_accountset.html',{'dev': dev,"usr":usr})
    #----------------------------------------------------------------------------------------------------TeleCaller

def tc_base(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    context={
        "usr":usr,
        "count":count,
    }
    return render(request, 'telecaller/tc_base.html',context)

def tc_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    context={
        "usr":usr,
        "count":count,
    }
    return render(request, 'telecaller/tc_profile.html',context)

def tc_dashboard(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    count = All_leads.objects.filter(
    Q(telecaller_id=ids, followup_dt=date.today()) &
    (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry'))).count()
    print(count)
    dt=date.today()
    context={
        "usr":usr,
        "dt":dt,
        "count":count
    }
    return render(request, 'telecaller/tc_dashboard.html',context)

def tc_accountset(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.filter(id=callerid)
    return render(request, 'telecaller/tc_accountset.html', {'caller': caller,"usr":usr})

def tc_change_pass(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')

    caller = user_registration.objects.filter(id=callerid)

    if request.method == 'POST':
        abc = user_registration.objects.get(id=callerid)
        cur = abc.password
        oldps = request.POST["currentPassword"]
        newps = request.POST["newPassword"]
        cmps = request.POST["confirmPassword"]
        if oldps == cur:
            if oldps != newps:
                if newps == cmps:
                    abc.password = request.POST.get('confirmPassword')
                    abc.save()
                    return render(request, 'telecaller/tc_ch_pass.html', {'caller': caller,"usr":usr})
            elif oldps == newps:
                messages.add_message(request, messages.INFO, 'Current and New password same')
            else:
                messages.info(request, 'Incorrect password same')

            return render(request, 'telecaller/tc_ch_pass.html', {'caller': caller,"usr":usr})
        else:
            messages.add_message(request, messages.INFO, 'old password wrong')
            return render(request, 'telecaller/tc_ch_pass.html', {'caller': caller,"usr":usr})
    return render(request, 'telecaller/tc_ch_pass.html', {'caller': caller,"usr":usr})

def tc_imagechange(request, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        devid = request.session['userid']
    else:
        return redirect('/')
    dev = user_registration.objects.filter(id=devid)
    if request.method == 'POST':
        abc = user_registration.objects.get(id=id)
        abc.photo = request.FILES['filename']
        
        abc.save()
        return redirect('tc_accountset')
    return render(request, 'telecaller/tc_accountset.html',{'dev': dev,"usr":usr})




def tc_view_leads(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.filter(id=callerid)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_view_leads.html', {'caller': caller,"usr":usr,"count":count})

def tc_view_current_leads(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.get(id=callerid)
    status_values=['no',' ','Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']
    cur_leads = All_leads.objects.filter(telecaller_id=ids,status__in=status_values).order_by('-assign_dt')
            

    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()

    return render(request, 'telecaller/tc_view_current_leads.html', {'caller': caller,"usr":usr,'cur_leads':cur_leads,"count":count})


def tc_update_status(request):
    ids=request.session['userid']
    id = request.GET.get('id')
    value = request.GET.get('value')
    lead = All_leads.objects.get(id = id)
    
    lead.status = value
    lead.followup_dt=None
    lead.save()
    print(lead.status)
    return JsonResponse({"status": " not",'value': lead.status})

def tc_update_followup_dt(request):
    ids=request.session['userid']
    id = request.GET.get('id')
    value = request.GET.get('value')
    lead = All_leads.objects.get(id = id)
    
    lead.followup_dt=value
    lead.save()
    print(lead.status)
    return JsonResponse({"status": " not",'value': lead.status})


def tc_view_notifications(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.get(id=callerid)
    # cur_leads = All_leads.objects.filter(telecaller_id=ids,followup_dt=date.today())
    cur_leads = All_leads.objects.filter(
    Q(telecaller_id=ids, followup_dt=date.today()) &
    (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))
)
    # print('hi')
    # for i in cur_leads:
    #     print(i)

    return render(request, 'telecaller/tc_view_notifications.html', {'caller': caller,"usr":usr,'cur_leads':cur_leads})



def tc_leave_home(request):
    if 'userid' in request.session:
        
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
     
        return render(request, 'telecaller/tc_leave_home.html',{'usr':usr,'count':count})
    else:
        return redirect('/')

def tc_leave_aply(request):
    if 'userid' in request.session:
        ids=request.session['userid']
        usr = user_registration.objects.get(id=ids)
        return render(request, 'telecaller/tc_leave_apply.html',{'usr':usr})
    else:
        return redirect('/')

def tc_leave_form(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        des1 = user_registration.objects.get(id=ids)
        if request.method == "POST":
            leaves = leave()
            leaves.from_date = request.POST['from']
            leaves.to_date = request.POST['to']
            leaves.leave_status = request.POST['haful']
            leaves.reason = request.POST['reason']
            leaves.user_id = ids
            leaves.status = "submitted"
            leaves.designation_id = des1.id
            leaves.leaveapprovedstatus=0
            
            start = datetime.strptime(leaves.from_date, '%Y-%m-%d').date() 
            end = datetime.strptime(leaves.to_date, '%Y-%m-%d').date()

            diff = (end  - start).days
    
            leaves.save()
        return  redirect('tc_all_leave')  
    else:
        return redirect('/')

def tc_all_leave(request):
    if 'userid' in request.session:
        if 'userid' in request.session:
            ids=request.session['userid']
            usr = user_registration.objects.get(id=ids)
        else:
           return redirect('/')
        mem = user_registration.objects.filter(id=ids)
        var = leave.objects.filter(user_id=ids)
        return render(request, 'telecaller/tc_all_leave.html',{'var': var,'usr':usr})
    else:
        return redirect('/')



# ------------------------------------------<Alen Antony> <Datamanager>-------------------------------------

def assigned_persons(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assign=All_leads.objects.filter(data_manager_id=ids).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')
    telecaller=user_registration.objects.filter(department='Telecaller')
    print(assign)
    context={
        "usr":usr,
        'assign':assign,
        'telecaller':telecaller,       

    }
    return render(request,'dm_manager/dm_assigned_persons.html', context)


def assigned_person_details(request,id,pk):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    assign=All_leads.objects.filter(data_manager_id=ids,telecaller_id=id,assign_dt=pk).order_by('-assign_dt')
    print(assign)
    context={
        "usr":usr,
        'assign':assign


    }
    return render(request,'dm_manager/dm_assigned-person_det.html', context)    











# ------------------------------------------<Alen Antony> <Telecaller>-------------------------------------

def tc_follow(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today = date.today()
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']
    count=All_leads.objects.filter(telecaller_id=ids,status__in=status_values,followup_dt=today).count()


    context={
        "usr":usr,
        'count' :count
             

    }

    return render(request,'telecaller/tc_follow.html',context)



def tc_followup(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Will contact after enquiry','Will contact if interested','Follow up']
    followup=All_leads.objects.filter(telecaller_id=ids,status__in=status_values).values('followup_dt').annotate(total_leads=Count('id')).order_by('-followup_dt')
    
    context={
        "usr":usr, 
        'followup':followup,      

    }
    return render(request,'telecaller/tc_followup_datewise.html',context)

def tc_followup_det(request,pk):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Will contact after enquiry','Will contact if interested','Follow up']

    followup=All_leads.objects.filter(telecaller_id=ids,followup_dt=pk,status__in=status_values).order_by('-followup_dt')

    context={
        "usr":usr,
        'followup':followup
       

    }
    return render(request,'telecaller/tc_followup.html',context)


def tc_close(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today = date.today()

    # status_values=['Not interested','Already contacted', 'Already joined in Altos','Already joined for another training', 'Not interested for payment','Recently join the job','Already done another internship recently','looking for direct job','payment is not affordable','Payment issue','Closed','Rejected']
    status_values=['Closed']
    count=All_leads.objects.filter(telecaller_id=ids,status__in=status_values,assign_dt=today).count()

    context={
        "usr":usr,
        'count':count, 
            

    }
    return render(request,'telecaller/tc_close.html',context)


def tc_closed(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Closed']
    # closed=All_leads.objects.filter(telecaller_id=ids,status__in=status_values).order_by('-assign_dt')
    closed=All_leads.objects.filter(telecaller_id=ids,status__in=status_values).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    context={
        "usr":usr, 
        'closed':closed,      

    }
    return render(request,'telecaller/tc_closed.html',context)    

def tc_flt_day_closed(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    day_field=request.POST.get('day')
    
    # status_values=['Not interested','Already contacted', 'Already joined in Altos','Already joined for another training', 'Not interested for payment','Recently join the job','Already done another internship recently','looking for direct job','payment is not affordable','Payment issue','Closed','Rejected']
    status_values=['Closed']

    # closed=All_leads.objects.filter(telecaller_id=ids,status__in=status_values,assign_dt=day_field).order_by('-assign_dt')
    closed=All_leads.objects.filter(telecaller_id=ids,status__in=status_values,assign_dt=day_field).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')


    
    context={
        "usr":usr,
        'closed':closed,
    }
    return render(request, 'telecaller/tc_closed.html',context)


def tc_flt_month_closed(request):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

    # status_values=['Not interested','Already contacted', 'Already joined in Altos','Already joined for another training', 'Not interested for payment','Recently join the job','Already done another internship recently','looking for direct job','payment is not affordable','Payment issue','Closed','Rejected']
    status_values=['Closed']

    # closed=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number) & Q(status__in=status_values),telecaller_id=ids, ).order_by('-assign_dt')
    closed=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number) & Q(status__in=status_values),telecaller_id=ids,).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')
    
    
    
    context={
        "usr":usr,
        'closed':closed
       

    }
    return render(request, 'telecaller/tc_closed.html',context)

def tc_closed_det(request,pk):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Closed']

    closed=All_leads.objects.filter(telecaller_id=ids,assign_dt=pk,status__in=status_values).order_by('-assign_dt')

    context={
        "usr":usr,
        'closed':closed
       

    }
    return render(request,'telecaller/tc_closed_det.html',context)



def tc_filter_day_previous_leads(request):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    caller = user_registration.objects.get(id=callerid)
    # print(callerid)
    pre_leads = All_leads.objects.filter(telecaller_id=ids,assign_dt=day_field)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_view_previous_leads.html', {'caller': caller,"usr":usr,'pre_leads':pre_leads,"count":count})

def tc_filter_month_previous_leads(request):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year

    caller = user_registration.objects.get(id=callerid)
    # print(callerid)
    pre_leads = All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number),telecaller_id=ids)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_view_previous_leads.html', {'caller': caller,"usr":usr,'pre_leads':pre_leads,"count":count})

def get_cust_mail(request):

  
    cust = request.GET.get('cust')
    print(cust)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Will contact after enquiry','Will contact if interested','Follow up']
    item = All_leads.objects.filter(telecaller_id=cust, status__in=status_values,followup_dt=date.today()).count()
    email = item
    no_status = ["no"," "]
    no_call_cnt = All_leads.objects.filter(telecaller_id=cust, status__in=no_status).count()
    print(item)
    return JsonResponse({"status": " not", 'email': email,"no_call_cnt":no_call_cnt})
    return redirect('/')
#----------------------------------------------------------------------Sumayya-----DataManager-------------------------------------------
def dm_base(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
   
    context={
        "usr":usr,
    }
    return render(request, 'dm_manager/dm_ base.html',context)

def dm_telecallers(request):
    
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    telecallers = user_registration.objects.filter(department='Telecaller')
   
    return render(request, 'dm_manager/dm_telecallers.html', {'telecallers': telecallers,'usr':usr})

def dm_followup_datewise(request,tid):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    telecaller = user_registration.objects.get(id=tid)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']

    leads_by_date = All_leads.objects.filter(telecaller_id=tid,status__in=status_values).values('assign_dt').annotate(total_leads=Count('id'))

    
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        'tid': tid
    }
    return render(request, 'dm_manager/dm_followup_datewise.html', context)

def dm_follow_tele_det(request,tid,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Call you back','Will contact after enquiry','Will contact if interested','Follow up']
    assign=All_leads.objects.filter(status__in=status_values,telecaller_id=tid,assign_dt=date).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'assign':assign,
        "tid":tid


    }
    return render(request,'dm_manager/dm_follow_tele_det.html', context)
    



def dm_filter_day_followup(request,pk):
    print("day")
    print(pk)
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    day_field=request.POST.get('day')
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Will contact after enquiry','Will contact if interested','Follow up']
    followup=All_leads.objects.filter(telecaller_id = pk,status__in=status_values,followup_dt=day_field).order_by('-followup_dt')
    
    context={
        "usr":usr, 
        'followup':followup, 
        't_id': pk,

    }
    return render(request,'dm_manager/dm_followup.html',context)

def dm_filter_month_followup(request,pk):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)

    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year
    status_values=['Will inform','Interested','Want to visit office', 'Please send details through WhatsApp','Join later','Will contact after enquiry','Will contact if interested','Follow up']
    followup=All_leads.objects.filter(Q(followup_dt__year=selected_year_number) & Q(followup_dt__month=selected_month_number),telecaller_id = pk,status__in=status_values).order_by('-followup_dt')
    
    context={
        "usr":usr, 
        'followup':followup,
        't_id': pk,      

    }
    return render(request,'dm_manager/dm_followup.html',context)

def dm_telecallers_no_calling(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    telecallers = user_registration.objects.filter(department='Telecaller')
   
    return render(request, 'dm_manager/dm_telecallers_no_calling.html', {'telecallers': telecallers,'usr':usr})
    # ids=request.session['userid']
    # usr = user_registration.objects.get(id=ids)
    # if request.session.has_key('userid'):
    #     usr = request.session['userid']
    #     telecallers = user_registration.objects.filter(department='Telecaller')
    #     context={'telecallers': telecallers,
    #              'usr':usr}
    #     return render(request, 'dm_manager/dm_telecallers_no_calling.html',context)
    # else:
    #     return redirect('/')

def dm_no_calling(request,t_id):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(
    #     telecaller_id=t_id, status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(telecaller_id=t_id,status__in=status_list).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    return render(request, 'dm_manager/dm_no_calling.html', {'leads': leads, 'usr': usr, 't_id':t_id})
    
def dm_filter_day_no_calling(request,t_id):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(assign_dt=day_field,
    #     telecaller_id=t_id, status__in=status_list).order_by('-assign_dt')

    leads=All_leads.objects.filter(telecaller_id=t_id,status__in=status_list,assign_dt=day_field).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')
    return render(request, 'dm_manager/dm_no_calling.html', {'leads': leads, 'usr': usr,'t_id':t_id,})

def dm_filter_month_no_calling(request,t_id):
    print('month')
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year
    status_list = ["no", " "]
    print('month1')
    # leads = All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number),
    #     telecaller_id=t_id, status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number) & Q(status__in=status_list),telecaller_id=t_id,).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')


    print('month')
    for i in leads:
        print(leads)
    
    return render(request, 'dm_manager/dm_no_calling.html', {'leads': leads, 'usr': usr,'t_id':t_id})

def dm_no_calling_det(request,t_id,pk,):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(
    #     telecaller_id=t_id, status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(telecaller_id=t_id,status__in=status_list,assign_dt=pk).order_by('-assign_dt')

    return render(request, 'dm_manager/dm_no_calling_det.html', {'leads': leads, 'usr': usr,})



def dm_executive_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executives = user_registration.objects.filter(department='Digital Marketing Executive')
   
    return render(request, 'dm_manager/dm_executive_home.html', {'executives': executives,'usr':usr})
    # ids=request.session['userid']
    # usr = user_registration.objects.get(id=ids)
    # if request.session.has_key('userid'):
    #     usr = request.session['userid']
        
    # else:
    #     return redirect('/')
    # executives = user_registration.objects.filter(department='Digital Marketing Executive')
    # return render(request, 'dm_manager/dm_executive_home.html',{'executives': executives,"usr":usr})

def dm_view_ex_clients(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    print(id)
    work_as=work_asign.objects.filter(exe_name=id).values('client_name_id').distinct()
    last=work_asign.objects.filter(exe_name=id).last()
    work=Work.objects.all()
    cl=client_information.objects.all()

   
    context={
        "usr":usr,
        "work_as":work_as,
        "work":work,
        "cl":cl,
        "last":last,
        "e_id": id
    }
    return render(request, 'dm_manager/dm_daily_work_clint.html',context)

def dm_daily_work_det(request,cid,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executive = user_registration.objects.get(id=eid)
    work_as=work_asign.objects.filter(exe_name=eid)
    works=Work.objects.filter(client_name_id=cid).order_by("-id")
    daily=daily_work.objects.filter(user=eid)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dts=date.today()
    dl_leeds=daily_leeds.objects.all()
    cr_date=date.today()
    
    context={
        "usr":usr,
        "cr_date":cr_date,
        "daily":daily,
        "work_as":work_as,
        "works":works,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,
        "dts":dts,
        "executive":executive
        
    }
    return render(request, 'dm_manager/dm_daily_work_det.html',context)

def dm_view_leads_home(request):
  
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executives = user_registration.objects.filter(department='Digital Marketing Executive')
   
    return render(request, 'dm_manager/dm_view_leads_home.html', {'executives': executives,'usr':usr})

def dm_leads_datewise(request,eid):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    exc = user_registration.objects.get(id=eid)
    # leads_by_date = All_leads.objects.values('date').annotate(total_leads=Count('id'))
    leads_by_date = All_leads.objects.filter(executive=exc).values('date').annotate(total_leads=Count('id'))

    for lead in leads_by_date:
        lead['target'] = random.randint(70, 120)
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        'eid': eid
    }
    return render(request, 'dm_manager/dm_leads_datewise.html', context)

def dm_view_all_leads(request, eid, date=None):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exc = user_registration.objects.get(id=eid)
    if date:
        leads = All_leads.objects.filter(date=date,executive=exc)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        'eid': eid,
        'date': date
    }
    return render(request, 'dm_manager/dm_view_all_leads.html', context)

def dm_filter_status_leads(request, eid, date=None):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    exc = user_registration.objects.get(id=eid)
    status_value=request.POST.get('status')
    print(status_value)
    if status_value == 'Non calling':
        status_list = ["no", " "]
    else:
        status_list =[status_value]

    print(status_list)
    if date:
        leads = All_leads.objects.filter(date=date,executive=exc,status__in=status_list)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        'eid': eid,
        'date': date
    }
    return render(request, 'dm_manager/dm_view_all_leads.html', context)

def dm_ad_delay_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_task=lead_delay.objects.all()
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "exe":exe
    }
    return render(request, 'dm_manager/dm_ad_delay.html', context)


def dm_ad_delay_flt(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    if request.method=="POST":
        exe=request.POST.get('exe_name')
        start_date=request.POST.get('start_date')
        end_date=request.POST.get('end_date')
        w_mails = dm_warning_mails.objects.filter(executive_id=exe)

        st_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        # dl_task = lead_delay.objects.filter(date__gte=st_dt, date__lte=end_dt,executive_id=exe).order_by('-date')

        wk_assign = work_asign.objects.filter(exe_name_id=exe)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,start_date__gte=st_dt,end_date__lte=end_dt,status__in=status_values).order_by('-start_date')
        print(work)
        for i in work:
            i.difference = i.target - 10

        today = date.today()
        print(today)
        context = {
            "usr":usr,
            # 'dl_task': dl_task,
            "exe":exes,
            # 'warning':warning,
            'eid':exe,
            'work':work,
            'today':today,
            'w_mails':w_mails

        }
    
        return render(request,'dm_manager/dm_ad_delay_det.html', context)
    return redirect("dm_ad_delay_det")



# --------------------------------------------------------Sumayya----telecaller-----------------------------------------------------------------


def tc_no_calling(request):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(
    #     telecaller_id=ids, status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(telecaller_id=ids,status__in=status_list).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    return render(request, 'telecaller/tc_no_calling.html', {'leads': leads, 'usr': usr})



def tc_filter_day_no_calling(request):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    day_field=request.POST.get('day')
    print(day_field)
    caller = user_registration.objects.get(id=callerid)
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(telecaller_id=ids,assign_dt=day_field,status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(telecaller_id=ids,status__in=status_list,assign_dt=day_field).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_no_calling.html', {'caller': caller,"usr":usr,'leads':leads,"count":count})


def tc_filter_month_no_calling(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    
    month_field=request.POST.get('month')
    print(month_field)
    selected_month_date = datetime.strptime(month_field, '%Y-%m')

    selected_month_number = selected_month_date.month
    selected_year_number = selected_month_date.year
    print(selected_year_number)
    print(selected_month_number)
    caller = user_registration.objects.get(id=callerid)
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number) & Q(status__in=status_list),telecaller_id=ids ).order_by('-assign_dt')
    leads=All_leads.objects.filter(Q(assign_dt__year=selected_year_number) & Q(assign_dt__month=selected_month_number) & Q(status__in=status_list),telecaller_id=ids,).values('telecaller_id','assign_dt').distinct().annotate(count=Count('telecaller_id')).order_by('-assign_dt')

    for i in leads:
        print(i)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_no_calling.html', {'caller': caller,"usr":usr,'leads':leads,"count":count})

def tc_no_calling_det(request,pk):
    if 'userid' in request.session:
        ids = request.session['userid']
        usr = user_registration.objects.get(id=ids)
        print(ids)
    else:
        return redirect('/')
    status_list = ["no", " "]
    # leads = All_leads.objects.filter(
    #     telecaller_id=ids, status__in=status_list).order_by('-assign_dt')
    leads=All_leads.objects.filter(telecaller_id=ids,assign_dt=pk,status__in=status_list).order_by('-assign_dt')

    return render(request, 'telecaller/tc_no_calling_det.html', {'leads': leads, 'usr': usr})


def tc_previous_leads_datewise(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    followup = ["Interested","Want to visit office","Will inform","Will contact if interested","Join later","Call you back",
                "Will contact after enquiry","Follow up"]
    non_calling = ["no"," "]
    leads_by_date = All_leads.objects.filter(telecaller_id=ids).values('assign_dt').annotate(total_leads=Count('id'), non_calling=Count('id', filter=Q(status__in=non_calling)),
                                                              followup=Count('id', filter=Q(status__in=followup)))
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    
    context = {
         "usr":usr,
        'leads_by_date': leads_by_date,
        'count':count
    }
    return render(request, 'telecaller/tc_previous_leads_datewise.html', context)

def tc_view_previous_leads(request,dt=None):
    ids=request.session['userid']
    print(ids)
    usr = user_registration.objects.get(id=ids)
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
    caller = user_registration.objects.get(id=callerid)
    # print(callerid)
    pre_leads = All_leads.objects.filter(telecaller_id=ids,assign_dt=dt)
    count = All_leads.objects.filter(
            Q(telecaller_id=ids, followup_dt=date.today()) &
            (Q(status='Interested') | Q(status='Want to visit office') | Q(status='Will inform') | Q(status='Will contact if interested') | 
            Q(status='Join later') | Q(status='Call you back') | Q(status='Will contact after enquiry') | Q(status='Follow up'))).count()
    return render(request, 'telecaller/tc_view_previous_leads.html', {'caller': caller,"usr":usr,'pre_leads':pre_leads,"count":count})


# --------------------------------------------------------Alen Antony----DMhead-----------------------------------------------------------------

def he_delay_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_task=lead_delay.objects.all()
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "dl_task":dl_task,
        "exe":exe
    }
    return render(request, 'head/he_delay_det.html', context)


def he_delay_flt(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    if request.method=="POST":
        exe=request.POST.get('exe_name')
        start_date=request.POST.get('start_date')
        end_date=request.POST.get('end_date')
        w_mails = dm_warning_mails.objects.filter(executive_id=exe)

        st_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        # dl_task = lead_delay.objects.filter(date__gte=st_dt, date__lte=end_dt,executive_id=exe).order_by('-date')

        wk_assign = work_asign.objects.filter(exe_name_id=exe)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,start_date__gte=st_dt,end_date__lte=end_dt,status__in=status_values).order_by('-start_date')
        print(work)
        for i in work:
            i.difference = i.target - 10

        today = date.today()
        print(today)
        context = {
            "usr":usr,
            # 'dl_task': dl_task,
            "exe":exes,
            # 'warning':warning,
            'eid':exe,
            'work':work,
            'today':today,
            'w_mails':w_mails

        }
    
        return render(request,'head/he_delay_flt.html', context)
    return redirect("he_delay_det")

def he_warning_mail(request,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        desc = request.POST.get('description')
        file = request.FILES['file']
        wid = request.POST.get('wid')
        print(wid)
        data = dm_warning_mails(executive_id=eid,work_id=wid,description=desc,date=c_date,file=file)
        data.save()
        work = Work.objects.get(id=wid)
        work.mail_status='yes'
        work.save()

    return redirect("he_delay_det")


def ad_delay_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dm_head = user_registration.objects.filter(department='Digital Marketing Head')
    status_values = ["no"," "]
    delay_count = Work.objects.filter(status__in=status_values).count()
    # delay_count = lead_delay.objects.filter(status='no').count()
    return render(request, 'admin/ad_delay_home.html', {'dm_head': dm_head,'usr':usr,'delay_count':delay_count})
    
    
def ad_delay(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    w_mails = dm_warning_mails.objects.all()
    status_values = ["no"," "]
    work = Work.objects.filter(status__in=status_values).order_by('-start_date')
    for i in work:
        if i.target is not None:
            i.difference = i.target - 10
        else:
            i.difference = None
 
    
    today = date.today()
    context = {
        "usr":usr,
        "exe":exes,
        'work':work,
        'today':today,
        'w_mails':w_mails

    }

    return render(request,'admin/ad_delay.html', context)
    
    

def ad_warning_mail1(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        desc = request.POST.get('description')
        file = request.FILES['file']
        wid = request.POST.get('wid')
        eid = work_asign.objects.get(work_id=wid).exe_name_id
        print(wid)
        print(eid)
        data = dm_warning_mails(executive_id=eid,work_id=wid,description=desc,date=c_date,file=file)
        data.save()

    return redirect("ad_delay")



def ad_warning_mail(request,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        desc = request.POST.get('description')
        file = request.FILES['file']
        wid = request.POST.get('wid')
        print(wid)
        data = dm_warning_mails(executive_id=eid,work_id=wid,description=desc,date=c_date,file=file)
        data.save()

    return redirect("ad_delay")


def ad_delay_fltr(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    if request.method=="POST":
        exe=request.POST.get('exe_name')
        start_date=request.POST.get('start_date')
        end_date=request.POST.get('end_date')
        w_mails = dm_warning_mails.objects.filter(executive_id=exe)

        st_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        # dl_task = lead_delay.objects.filter(date__gte=st_dt, date__lte=end_dt,executive_id=exe).order_by('-date')

        wk_assign = work_asign.objects.filter(exe_name_id=exe)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,start_date__gte=st_dt,end_date__lte=end_dt,status__in=status_values).order_by('-start_date')
        print(work)
        for i in work:
            if i.target is not None:
                i.difference = i.target - 10
            else:
                i.difference = None 

        today = date.today()
        context = {
            "usr":usr,
            # 'dl_task': dl_task,
            "exe":exes,
            # 'warning':warning,
            'eid':exe,
            'work':work,
            'today':today,
            'w_mails':w_mails

        }
    
        return render(request,'admin/ad_delay_flt.html', context)
    return redirect("ad_delay")
    
    
#-----------------------------------------------------------------------Sumayya-----DM Head------------------------------------------------------

def he_waste_data(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads_by_date = All_leads.objects.filter(status__in=status_values).values('assign_dt').annotate(total_leads=Count('id')).order_by('-assign_dt')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'head/he_waste_data.html', context)

def he_waste_data_det(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads=All_leads.objects.filter(status__in=status_values,assign_dt=date).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'head/he_waste_data_det.html', context)

def he_executive_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executives = user_registration.objects.filter(department='Digital Marketing Executive')
   
    return render(request, 'head/he_executive_home.html', {'executives': executives,'usr':usr})
   

def he_view_ex_clients(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    print(id)
    work_as=work_asign.objects.filter(exe_name=id).values('client_name_id').distinct()
    last=work_asign.objects.filter(exe_name=id).last()
    work=Work.objects.all()
    cl=client_information.objects.all()

   
    context={
        "usr":usr,
        "work_as":work_as,
        "work":work,
        "cl":cl,
        "last":last,
        "e_id": id
    }
    return render(request, 'head/he_daily_work_clint.html',context)

def he_daily_work_det(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    work=he_daily_work.objects.filter(user=ids,cl_id=client.id)
    dts=date.today()
    cr_date=date.today()
    

    context={
        "usr":usr,
        'client':client,
        "cr_date":cr_date,
        "dts":dts,
        "work":work,
        
        
    }
    return render(request,'head/he_daily_work_det.html',context)  
  

def he_change_status(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        a_count = request.POST.get('ach_target')
        wid = request.POST.get('wid')
        print(wid)
        w = Work.objects.get(id=wid)
        w.achived = a_count
        w.save()
        w_as = work_asign.objects.get(work_id=wid)
        eid = w_as.exe_name_id
        cid = w_as.client_name_id
        
        

    return redirect("he_daily_work_details", cid=cid,eid=eid)
#-----------------------------

def he_all_leads_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'head/he_all_leads_exe.html',context)

def he_view_exe_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads_by_date = All_leads.objects.filter(executive=id).values('date').annotate(total_leads=Count('id'))
    for lead in leads_by_date:
        lead['target'] = random.randint(70, 120)
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        "id":id
    }
    return render(request, 'head/he_view_exe_det.html', context)

def he_view_all_leads(request, date, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if date:
        leads = All_leads.objects.filter(date=date, executive=id)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        "id":id,
        "date":date
    }
    return render(request, 'head/he_view_all_leads.html', context)


#-----------------------------------------------------------------------Sumayya-----Admin------------------------------------------------------

def ad_waste_data(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads_by_date = All_leads.objects.filter(status__in=status_values).values('assign_dt').annotate(total_leads=Count('id')).order_by('-assign_dt')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'admin/ad_waste_data.html', context)

def ad_waste_data_det(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads=All_leads.objects.filter(status__in=status_values,assign_dt=date).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'admin/ad_waste_data_det.html', context)

#-----------------------------------------------------------------------Sumayya-----Data Manager------------------------------------------------------

def dm_waste_data(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads_by_date = All_leads.objects.filter(status__in=status_values).values('assign_dt').annotate(total_leads=Count('id')).order_by('-assign_dt')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'dm_manager/dm_waste_data.html', context)

def dm_waste_data_det(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    status_values = ["Rejected","Not interested","Will contact if interested","Already contacted","Already joined in Altos",
                     "Already joined for another training","Not interested for payment","Recently join the job",
                     "Already done another internship recently","looking for direct job","payment is not affordable"]
    leads=All_leads.objects.filter(status__in=status_values,assign_dt=date).order_by('-assign_dt')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'dm_manager/dm_waste_data_det.html', context)
    
    
# --------------------------------------------------------Alen-----Admin-----------------------------------------------------------------


def ad_view_client(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)

    client=client_information.objects.all()

    context={
        "usr":usr,
        "client":client,
    }
    return render(request, 'admin/ad_view_client.html',context)

def ad_daily_work_head(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    work=he_daily_work.objects.filter(cl_id=client.id).order_by('-date')
    dts=date.today()
    cr_date=date.today()
    

    context={
        "usr":usr,
        'client':client,
        "cr_date":cr_date,
        "dts":dts,
        "work":work,
        
        
    }
    return render(request,'admin/ad_daily_work_head.html',context) 
    
    
# --------------------------------------------------------Alen-----DM Head-----------------------------------------------------------------

def he_create_work(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    return render(request, 'head/he_create_work.html',context)


def he_save_create_work(request):
    client = client_information()
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    if request.session.has_key('userid'):
        userid = request.session['userid']
    else:
       return redirect('/')
    if request.method == 'POST':
        client.client_name = request.POST.get('client_name')
        
        client.client_address = request.POST.get('client_address')
        client.client_mail = request.POST.get('client_mail')
        client.bs_name = request.POST.get('bs_name')
        client.bs_website = request.POST.get('bs_website',None)
       
        client.bs_location = request.POST.get('bs_location')
        client.client_files = request.FILES.get('client_files',None)
        client.seo = request.POST.get('seo',None)
        client.seo_txt = request.POST.get('seo_txt',None)
        client.seo_file = request.FILES.get('seo_file',None)
        client.seo_target = request.POST.get('seo_target',None)

        client.on_pg = request.POST.get('onpage',None)
        client.on_pg_txt = request.POST.get('on_txt',None)
        client.on_pg_file = request.FILES.get('on_file',None)
        client.on_pg_target = request.POST.get('on_target',None)

        client.off_pg = request.POST.get('offpage',None)
        client.off_pg_txt = request.POST.get('off_txt',None)
        client.off_pg_file = request.FILES.get('off_file',None)
        client.off_pg_target = request.POST.get('off_target',None)

        client.smm = request.POST.get('smm',None)
        client.smm_txt = request.POST.get('smm_txt',None)
        client.smm_file = request.FILES.get('smm_file',None)
        client.smm_target = request.POST.get('smm_target',None)
        client.smo = request.POST.get('smo',None)
        client.smo_txt = request.POST.get('smo_txt',None)
        client.smo_file = request.FILES.get('smo_file',None)
        client.smo_target = request.POST.get('smo_target',None)

        client.sem = request.POST.get('sem',None)
        client.sem_txt = request.POST.get('sem_txt',None)
        client.sem_file = request.FILES.get('sem_file',None)
        client.sem_target = request.POST.get('sem_target',None)
        client.em = request.POST.get('em',None)
        client.em_txt = request.POST.get('em_txt',None)
        client.em_file = request.FILES.get('em_file',None)
        client.em_target = request.POST.get('em_target',None)
        client.cm = request.POST.get('cm',None)
        client.cm_txt = request.POST.get('cm_txt',None)
        client.cm_file = request.FILES.get('cm_file',None)
        client.cm_target = request.POST.get('cm_target',None)
        client.am = request.POST.get('am',None)
        client.am_txt = request.POST.get('am_txt',None)
        client.am_file = request.FILES.get('am_file',None)
        client.am_target = request.POST.get('am_target',None)
        client.mm = request.POST.get('mm',None)
        client.mm_txt = request.POST.get('mm_txt',None)
        client.mm_file = request.FILES.get('mm_file',None)
        client.mm_target = request.POST.get('mm_target',None)
        client.vm = request.POST.get('vm',None)
        client.vm_txt = request.POST.get('vm_txt',None)
        client.vm_file = request.FILES.get('vm_file',None)
        client.vm_target = request.POST.get('vm_target',None)

        client.lc = request.POST.get('lc',None)
        client.lc_txt = request.POST.get('lc_txt',None)
        client.lc_file = request.FILES.get('lc_file',None)
        client.lc_target = request.POST.get('lc_target',None)

        client.user=usr
        client.save()
        
        client = client_information.objects.get(id=client.id)
        
        labels = request.POST.getlist('label[]')
        text =request.POST.getlist('dis[]')
        
        if len(labels)==len(text):
            mapped = zip(labels,text)
            mapped=list(mapped)
            for ele in mapped:
            
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
        else:
            pass

        labels2 = request.POST.getlist('label2[]')
        text2 =request.POST.getlist('dis2[]')
        
        if len(labels2)==len(text2):
            mappeds = zip(labels2,text2)
            mappeds=list(mappeds)
            for ele in mappeds:
            
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='business_details')
        else: 
            pass
          
        
        files_req =request.FILES.getlist('file_add[]') 
        label_req =request.POST.getlist('label_req[]')
        dis_req =request.POST.getlist('dis_req[]') 
        target =request.POST.getlist('target[]')

        
        if len(files_req)==len(label_req)==len(dis_req)==len(target):
            mapped2 = zip(label_req,dis_req,files_req,target)
            mapped2=list(mapped2)
         
            for ele in mapped2:
                created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],file=ele[2],target=ele[3],user=usr,client=client,section='requirments')

        msg_success = "Save Successfully"
        context={
            "usr":usr,
            "msg_success":msg_success,
        }
        return redirect("he_project") 
        
    return redirect("he_create_work")


def he_view_work(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.filter(user=ids)
    return render(request,'head/he_view_work.html',{'client':client,"usr":usr,})  

def he_view_clint(request,id):
    client=client_information.objects.get(id=id)
    addicl=addi_client_info.objects.filter(client=client.id,section='client_information')
    addibs=addi_client_info.objects.filter(client=client.id,section='business_details')
    addirq=addi_client_info.objects.filter(client=client.id,section='requirments')
    
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
        "client":client,
        "addicl":addicl,
        "addibs":addibs,
        "addirq":addirq,
    }
    return render(request, 'head/he_view_client.html',context)    

def he_update_client(request,id):
    client = client_information.objects.get(id=id)
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    
    if request.session.has_key('userid'):
        userid = request.session['userid']
    else:
       return redirect('/')
    if request.method == 'POST':
        client.client_name = request.POST.get('client_name')
        
        client.client_address = request.POST.get('client_address')
        client.client_mail = request.POST.get('client_mail')
        client.bs_name = request.POST.get('bs_name')
        client.bs_website = request.POST.get('bs_website',None)
       
        client.bs_location = request.POST.get('bs_location')
        if request.FILES.get('client_files',None) == None:
            client.client_files=client.client_files
        else:
            client.client_files = request.FILES.get('client_files',None)
      
        client.seo = request.POST.get('seo',None)
        client.seo_txt = request.POST.get('seo_txt',None)
        if request.FILES.get('seo_file',None) == None:
            client.seo_file=client.seo_file
        else:
            client.seo_file = request.FILES.get('seo_file',None)

        client.on_pg = request.POST.get('onpage',None)
        client.on_pg_txt = request.POST.get('on_txt',None)
        client.on_pg_target = request.POST.get('on_target',None)
        if request.FILES.get('on_file',None) == None:
            client.on_pg_file=client.on_pg_file
        else:
            client.on_pg_file = request.FILES.get('on_file',None)


        client.off_pg = request.POST.get('offpage',None)
        client.off_pg_txt = request.POST.get('off_txt',None)
        client.off_pg_target = request.POST.get('off_target',None)
        if request.FILES.get('off_file',None) == None:
            client.off_pg_file=client.off_pg_file
        else:
            client.off_pg_file = request.FILES.get('off_file',None)

   


        client.smm = request.POST.get('smm',None)
        client.smm_txt = request.POST.get('smm_txt',None)
        client.smm_target = request.POST.get('smm_target',None)
     
        if request.FILES.get('smm_file',None) == None:
            client.smm_file=client.smm_file
        else:
            client.smm_file = request.FILES.get('smm_file',None)

        client.smo = request.POST.get('smo',None)
        client.smo_txt = request.POST.get('smo_txt',None)
        client.smo_target = request.POST.get('smo_target',None)
     
        if request.FILES.get('smo_file',None) == None:
            client.smo_file=client.smo_file
        else:
            client.smo_file = request.FILES.get('smo_file',None)

        client.sem = request.POST.get('sem',None)
        client.sem_txt = request.POST.get('sem_txt',None)
        client.sem_target = request.POST.get('sem_target',None)
    

        if request.FILES.get('sem_file',None) == None:
            client.sem_file=client.sem_file
        else:
            client.sem_file = request.FILES.get('sem_file',None)


        client.em = request.POST.get('em',None)
        client.em_txt = request.POST.get('em_txt',None)
        client.em_target = request.POST.get('em_target',None)
        if request.FILES.get('em_file',None) == None:
            client.em_file=client.em_file
        else:
            client.em_file = request.FILES.get('em_file',None)


        client.cm = request.POST.get('cm',None)
        client.cm_txt = request.POST.get('cm_txt',None)
        client.cm_target = request.POST.get('cm_target',None)

        if request.FILES.get('cm_file',None) == None:
            client.cm_file=client.cm_file
        else:
            client.cm_file = request.FILES.get('cm_file',None)


        client.am = request.POST.get('am',None)
        client.am_txt = request.POST.get('am_txt',None)
        client.am_target = request.POST.get('am_target',None)
        if request.FILES.get('am_file',None) == None:
            client.am_file=client.am_file
        else:
            client.am_file = request.FILES.get('am_file',None)


        client.mm = request.POST.get('mm',None)
        client.mm_txt = request.POST.get('mm_txt',None)
        client.mm_target = request.POST.get('mm_target',None)
        if request.FILES.get('mm_file',None) == None:
            client.mm_file=client.mm_file
        else:
            client.mm_file = request.FILES.get('mm_file',None)


        client.vm = request.POST.get('vm',None)
        client.vm_txt = request.POST.get('vm_txt',None)
        client.vm_target = request.POST.get('vm_target',None)
        if request.FILES.get('vm_file',None) == None:
            client.vm_file=client.vm_file
        else:
            client.vm_file = request.FILES.get('vm_file',None)

        client.lc = request.POST.get('lc',None)
        
        client.lc_txt = request.POST.get('lc_txt',None)
     
        client.lc_target = request.POST.get('lc_target',None)
        if request.FILES.get('lc_file',None) == None:
            client.lc_file=client.lc_file
        else:
            client.lc_file = request.FILES.get('lc_file',None)


        client.user=usr
        client.save()
        client = client_information.objects.get(id=id)
       

        
        client = client_information.objects.get(id=client.id)
        
        labels = request.POST.getlist('label[]')
        text =request.POST.getlist('dis[]')
        
        if len(labels)==len(text):
            mapped = zip(labels,text)
            mapped=list(mapped)
            count = addi_client_info.objects.filter(client=id,section='client_information').count()
            lb_count=len(labels)
            print(lb_count)
            for ele in mapped:
                
                    try:
                        adiclient = addi_client_info.objects.get(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1]))
                        
                    
                        if ((adiclient.labels==ele[0]) or (adiclient.discription==ele[1])):
                            created = addi_client_info.objects.filter(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1])).update(labels=ele[0],discription=ele[1])
                        
                        
                        elif ((adiclient.labels!=ele[0]) or (adiclient.discription!=ele[1])):
                            created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
                    
                        else:
                            pass
                            
                    except:
                        created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],user=usr,client=client,section='client_information')
                




        else:
            
            pass

        labels2 = request.POST.getlist('label2[]')
        text2 =request.POST.getlist('dis2[]')
        
        if len(labels2)==len(text2):
            mappeds = zip(labels2,text2)
            mappeds=list(mappeds)

      
            for ele in mappeds:
                try:
                    adiclient=addi_client_info.objects.get(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1]))
                    if ((adiclient.labels==ele[0]) or (adiclient.discription==ele[1])):
                        created = addi_client_info.objects.filter(Q(client=client),Q(labels=ele[0])|Q(discription=ele[1])).update(labels=ele[0],discription=ele[1])
                    elif ((adiclient.labels!=ele[0]) or (adiclient.discription!=ele[1])):
                        created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],client=client,user=usr,section='business_details')
                    else:
                        pass
                except:
                    created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],client=client,user=usr,section='business_details')

        else: 
            pass
        
      
          

        
        label_req =request.POST.getlist('label_req[]')
        dis_req =request.POST.getlist('dis_req[]')
        target =request.POST.getlist('target[]')
       
        if request.FILES.getlist('file_add[]') == []:
            img=addi_client_info.objects.filter(client=id,section='Requirments')
            files_req=[]
            for i in img:
                files_req.append(i.file)
       
           
        elif len(request.FILES.getlist('file_add[]')) != len(label_req):
          
            img=addi_client_info.objects.filter(client=id,section='Requirments')
            fr=[]
            for i in img:
                fr.append(i.file)
            fr2 =request.FILES.getlist('file_add[]') 
            files_req=fr+fr2
           
        else:
           
            files_req=request.FILES.getlist('file_add[]')

        
        
        if len(label_req)==len(dis_req)==len(target)==len(files_req):
            
            mapped2 = zip(label_req,dis_req,files_req,target)
            mapped2=list(mapped2)
           
            
            abs=addi_client_info.objects.filter(client=id,section='Requirments').delete()
            for ele in mapped2:
                    

                    created = addi_client_info.objects.get_or_create(labels=ele[0],discription=ele[1],file=ele[2],target=ele[3],user=usr,client=client,section='Requirments')
                
        else:
       
            pass

        msg_success = "Save Successfully"
        return redirect('he_view_clint',id)
    return redirect('he_view_clint',id)



def he_daily_work_client(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.all()
    return render(request,'head/he_daily_work_client.html',{'client':client,"usr":usr,})
    
    
def he_daily_work_done(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    if request.method == 'POST':
        daily= he_daily_work()
        daily.user=usr 
        daily.cl_id=client.id
        daily.date=date.today()
        daily.workdone =request.POST.get('workdone',None)
        dct_file = dict(request.FILES)
        lst_screenshot = dct_file['filed']
        lst_file = []
        for ins_screenshot in lst_screenshot:
            str_img_path = ""
            if ins_screenshot:
                img_emp = ins_screenshot
                fs = FileSystemStorage(location=settings.MEDIA_ROOT,base_url=settings.MEDIA_URL)
                str_img = fs.save(''.join(filter(str.isalnum, str(img_emp))), img_emp)
                str_img_path = fs.url(''.join(filter(str.isalnum, str_img)))
                lst_file.append('/media/'+''.join(filter(str.isalnum, str(img_emp))))
                daily.json_testerscreenshot = lst_file
        daily.save()
        return redirect('he_daily_work_det', pk=client.id)
        
        
#-----------------------------------------------------------------------Sumayya-----DM Head------------------------------------------------------

def he_waste_data_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'head/he_waste_data_home.html', {'usr':usr})

def he_waste_data_exe(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    leads_by_date = daily_leeds_exists.objects.values('date').annotate(total_leads=Count('id')).order_by('-date')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'head/he_waste_data_exe.html', context)

def he_waste_data_det_exe(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads=daily_leeds_exists.objects.filter(date=date).order_by('-date')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'head/he_waste_data_det_exe.html', context)
    
    
#-----------------------------------------------------------------------Sumayya-----Admin------------------------------------------------------

def ad_waste_data_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'admin/ad_waste_data_home.html', {'usr':usr})

def ad_waste_data_exe(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    leads_by_date = daily_leeds_exists.objects.values('date').annotate(total_leads=Count('id')).order_by('-date')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'admin/ad_waste_data_exe.html', context)

def ad_waste_data_det_exe(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads=daily_leeds_exists.objects.filter(date=date).order_by('-date')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'admin/ad_waste_data_det_exe.html', context)
    
    
#-----------------------------------------------------------------------Sumayya-----Data Manager------------------------------------------------------

def dm_waste_data_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'dm_manager/dm_waste_data_home.html', {'usr':usr})

def dm_waste_data_exe(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    leads_by_date = daily_leeds_exists.objects.values('date').annotate(total_leads=Count('id')).order_by('-date')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'dm_manager/dm_waste_data_exe.html', context)

def dm_waste_data_det_exe(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads=daily_leeds_exists.objects.filter(date=date).order_by('-date')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'dm_manager/dm_waste_data_det_exe.html', context)
    
    
def he_daily_work_det(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    work=he_daily_work.objects.filter(user=ids,cl_id=client.id)
    dts=date.today()
    cr_date=date.today()
    

    context={
        "usr":usr,
        'client':client,
        "cr_date":cr_date,
        "dts":dts,
        "work":work,
        
        
    }
    return render(request,'head/he_daily_work_det.html',context)
    
    
def he_daily_work_details(request,cid,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executive = user_registration.objects.get(id=eid)
    work_as=work_asign.objects.filter(exe_name=eid)
    works=Work.objects.filter(client_name_id=cid).order_by("-id")
    daily=daily_work.objects.filter(user=eid)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dts=date.today()
    dl_leeds=daily_leeds.objects.all()
    cr_date=date.today()

    for l in works:
        l.difference = l.target - 10
    
    context={
        "usr":usr,
        "cr_date":cr_date,
        "daily":daily,
        "work_as":work_as,
        "works":works,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,
        "dts":dts,
        "executive":executive
        
    }
    return render(request, 'head/he_daily_work_details.html',context)

#-----------------------------------------------------------------sumayya--tl module-----------------------------------------------------------

def tl_profile(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    context={
        "usr":usr,
    }
    
    return render(request, 'team_lead/tl_profile.html',context)

def tl_project(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request,'team_lead/tl_project.html',{"usr":usr})

def tl_view_works(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.all()
    return render(request,'team_lead/tl_view_works.html',{'client':client,"usr":usr,})

def tl_work_asign(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=pk)
    exe=user_registration.objects.filter(department='Digital Marketing Executive')
    return render(request,'team_lead/tl_work_asign.html',{'client':client,'exe':exe,"usr":usr,})

def tl_work_add(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method == 'POST':
        task = request.POST.get('task')
        des = request.POST.get('des')
        sdate=request.POST.get('sdate')
        edate=request.POST.get('edate')
        file=request.FILES.get('file')
        sub_tsk = request.POST.get('sub_tsk')
        target=request.POST.get('trgt')
        client=client_information.objects.get(id=id)
        json_data = request.POST.get('array', '')
        array = json.loads(json_data)
        w=Work(task=task,description=des,start_date=sdate,end_date=edate,file_attached=file,cl_name=client.bs_name,client_name=client)
        w.save()

        if w.task=="SEO":
            w.file_2=client.seo_file
            w.save()
            if sub_tsk=="On page":
                w.sub_task="On page"
                w.sub_des=client.on_pg_txt
                w.sub_file=client.on_pg_file
                w.target=request.POST.get('ontrgt')
                w.save()
            if sub_tsk=="Off page":
                w.sub_task="Off page"
                w.sub_des=client.off_pg_txt
                w.sub_file=client.off_pg_file
                w.target=request.POST.get('offtrgt')
                w.save()

        if w.task=="SMM":
            w.file_2=client.smm_file
            w.target=target
            w.save()
        if w.task=="SEM/PPC":
            w.file_2=client.sem_file
            w.target=target
            w.save()
        if w.task=="Email Marketing":
            w.file_2=client.em_file
            w.target=target
            w.save()   
        if w.task=="Content Marketing":
            w.file_2=client.cm_file
            w.target=target
            w.save() 
        if w.task=="Affiliate Marketing":
            w.file_2=client.am_file
            w.target=target
            w.save()   
        if w.task=="Mobile marketing":
            w.file_2=client.mm_file
            w.target=target
            w.save()  
        if w.task=="Video Marketing":
            w.file_2=client.vm_file
            w.target=target
            w.save() 
        if w.task=="SMO":
            w.file_2=client.smo_file
            w.target=target
            w.save() 
        if w.task=="Leads Collection":
            w.file_2=client.lc_file
            w.target=target
            w.save()            
        w=Work.objects.latest('id')
        for i in array:
            b=user_registration.objects.get(department="Digital Marketing Executive",fullname=i)
            c=work_asign(work_id=w.id,exe_name_id=b.id,client_name_id=client.id)
            c.save()
        return HttpResponse({"message": "success"})

def tl_view_work_asign_client(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.all()
    return render(request,'team_lead/tl_view_work_asign_client.html',{'client':client,"usr":usr,})

def tl_view_work_asign_exe(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    client=client_information.objects.get(id=id)
    work=work_asign.objects.filter(client_name=client.id)
    return render(request,'team_lead/tl_view_work_asign_exe.html',{"usr":usr,"w":work})


def tl_daily_task(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today=date.today()
    work=daily_work.objects.filter(date=today)
    master=Work.objects.all()
    sub_work=daily_work_sub.objects.all()
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()

    dl_leeds=daily_leeds.objects.all()
    return render(request,'team_lead/tl_daily_task.html',{'work':work,"usr":usr,"sub":sub_work,"dl_sub":dl_sub,"dl_off":dl_off,"dl_leeds":dl_leeds,'master':master})

def tl_add_correction_daily(request,id):
    if request.method=='POST':
        cor=request.POST.get('cor')
        daily=daily_work.objects.get(id=id)
        print(daily.user_id)
        c=correction(description=cor,daily_id=daily.id,executive_id=daily.user_id,event=None)
        c.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def tl_add_status(request,id):
    if request.method=='POST':
        status=request.POST.get('status')   
        daily=daily_work.objects.get(id=id)
        daily.status=status
        daily.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def tl_export_excel(request,id):

    filtered_data = daily_leeds.objects.filter(daily=id)

    # Create an Excel workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Add column headers to the Excel sheet
    headers = ['No.',"name","email_id","ph_no","location","qualification","year_of_passout","collegename","internship","internship_institute","internship_topic","internship_start","internship_end","duration","fresher_experience","previous_experience","company_name","register for what",'Duration' 'Experiense','Internship']  # "Replace with your actual column names
    sheet.append(headers)

    # Add data rows to the Excel sheet
    count = 1
    for item in filtered_data:
        
        row = [count,item.name,item.email_id,item.ph_no,item.location,item.qualification,item.year_of_passout,item.collegename,item.internship,item.internship_institute,item.internship_topic,item.internship_start,item.internship_end,item.duration,item.fresher_experience,item.previous_experience,item.company_name,item.register,item.ex_duration] # Replace with your actual column names
        sheet.append(row)
        count+=1

    # Set the response headers for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filtered_data.xlsx'

    # Save the Excel workbook to the response
    workbook.save(response)

    return response
    
def tl_workprogress_executive(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    prgs=progress_report.objects.all()
    return render(request,'team_lead/tl_workprogress_executive.html',{'prgs':prgs,"usr":usr,})

def tl_progress_report(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    work=progress_report.objects.get(id=pk)
    try:
        prv_work=progress_report.objects.filter(work_id=work.id).order_by('-end_date')[0]
    except:
        prv_work=None
    return render(request,'team_lead/tl_progress_report.html',{'work':work,"usr":usr,"prv_work":prv_work})

def tl_flt_progress(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    st_dt=request.POST.get('str_dt')
    en_dt=request.POST.get('end_dt')
  
    pr_work=progress_report.objects.filter(start_date__gte=st_dt,start_date__lte=en_dt)
    context={
        "usr":usr,
        "prgs":pr_work

    }
    return render(request, 'team_lead/tl_workprogress_executive.html',context)


def tl_smo_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request, 'team_lead/tl_smo_exe.html',{'exe':exe,"usr":usr})

def tl_view_post(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    today = date.today()
    event=Events.objects.filter(executive=id, start__date=today)
    hr_list=["09:00:00","10:00:00","11:00:00","12:00:00","13:00:00","14:00:00","15:00:00","16:00:00"]
    start_values = [i.start.strftime("%H:%M:%S") for i in event]
    unschedule = [datetime.strptime(elem, '%H:%M:%S').strftime('%I:%M %p') for elem in hr_list if elem not in start_values]
    return render(request,'team_lead/tl_view_post.html',{"usr":usr,'evnt':event,'unsh':unschedule})

def tl_add_correction(request,id):
    if request.method=='POST':
        cor=request.POST.get('cor')
        event=Events.objects.get(id=id)
        c=correction(description=cor,executive_id=event.executive.id,event_id=event.id,daily=None)
        c.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def tl_add_event_status(request,id):
    if request.method=='POST':
        status=request.POST.get('status')   
        event=Events.objects.get(id=id)
        event.status=status
        event.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  


def tl_cor_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request, 'team_lead/tl_cor_exe.html',{'exe':exe,"usr":usr})

def tl_cor_exe_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    cor=correction.objects.filter(executive=id)
    

    context={
        "usr":usr,
        "warn":cor,
    }
    return render(request, 'team_lead/tl_cor_exe_det.html',context)

def tl_all_leads_exe(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    tls=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "tls":tls
      
    }
    return render(request, 'team_lead/tl_all_leads_exe.html',context)

def tl_view_exe_det(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads_by_date = All_leads.objects.filter(executive=id).values('date').annotate(total_leads=Count('id'))
    for lead in leads_by_date:
        lead['target'] = random.randint(70, 120)
    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
        "id":id
    }
    return render(request, 'team_lead/tl_view_exe_det.html', context)

def tl_view_all_leads(request, date, id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if date:
        leads = All_leads.objects.filter(date=date, executive=id)
    else:
        leads = All_leads.objects.none()  # Return an empty queryset if no date is provided
    context = {
        "usr":usr,
        'leads': leads,
        "id":id,
        "date":date
    }
    return render(request, 'team_lead/tl_view_all_leads.html', context)

def tl_executive_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executives = user_registration.objects.filter(department='Digital Marketing Executive')
   
    return render(request, 'team_lead/tl_executive_home.html', {'executives': executives,'usr':usr})
   
def tl_view_ex_clients(request,id):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    print(id)
    work_as=work_asign.objects.filter(exe_name=id).values('client_name_id').distinct()
    last=work_asign.objects.filter(exe_name=id).last()
    work=Work.objects.all()
    cl=client_information.objects.all()

   
    context={
        "usr":usr,
        "work_as":work_as,
        "work":work,
        "cl":cl,
        "last":last,
        "e_id": id
    }
    return render(request, 'team_lead/tl_daily_work_clint.html',context)

def tl_daily_work_details(request,cid,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    executive = user_registration.objects.get(id=eid)
    work_as=work_asign.objects.filter(exe_name=eid)
    works=Work.objects.filter(client_name_id=cid).order_by("-id")
    daily=daily_work.objects.filter(user=eid)
    dl_sub=daily_work_sub.objects.all() 
    dl_off=daily_off_sub.objects.all()
    dts=date.today()
    dl_leeds=daily_leeds.objects.all()
    cr_date=date.today()

    for l in works:
        l.difference = l.target - 10
    
    context={
        "usr":usr,
        "cr_date":cr_date,
        "daily":daily,
        "work_as":work_as,
        "works":works,
        "dl_sub":dl_sub,
        "dl_off":dl_off,
        "dl_leeds":dl_leeds,
        "dts":dts,
        "executive":executive
        
    }
    return render(request, 'team_lead/tl_daily_work_details.html',context)
  

def tl_change_status(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        a_count = request.POST.get('ach_target')
        wid = request.POST.get('wid')
        print(wid)
        w = Work.objects.get(id=wid)
        w.achived = a_count
        w.save()
        w_as = work_asign.objects.get(work_id=wid)
        eid = w_as.exe_name_id
        cid = w_as.client_name_id
        
        

    return redirect("tl_daily_work_details", cid=cid,eid=eid)


def tl_feedback(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    return render(request,'team_lead/tl_feedback.html',{'exe':exe,"usr":usr,})


def tl_feedbacke1(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exe=user_registration.objects.get(id=pk)
    wrng=Warning.objects.filter(executive_id=exe.id)
    return render(request,'team_lead/tl_feedback1.html',{'exe':exe,'wrng':wrng,"usr":usr,})

    
def tl_feedback_submit(request,pk):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    if request.method=='POST':
        des=request.POST['des']
        typ=request.POST['option']
        warning=Warning(executive_id=pk,description=des,type=typ,)
        warning.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def tl_delay_det(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    dl_task=lead_delay.objects.all()
    exe=user_registration.objects.filter(department="Digital Marketing Executive")
    context={
        "usr":usr,
        "dl_task":dl_task,
        "exe":exe
    }
    return render(request, 'team_lead/tl_delay_det.html', context)


def tl_delay_flt(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    exes=user_registration.objects.filter(department="Digital Marketing Executive")
    if request.session.has_key('userid'):
        callerid = request.session['userid']
    else:
        return redirect('/')
   
    if request.method=="POST":
        exe=request.POST.get('exe_name')
        start_date=request.POST.get('start_date')
        end_date=request.POST.get('end_date')
        w_mails = dm_warning_mails.objects.filter(executive_id=exe)

        st_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        # dl_task = lead_delay.objects.filter(date__gte=st_dt, date__lte=end_dt,executive_id=exe).order_by('-date')

        wk_assign = work_asign.objects.filter(exe_name_id=exe)
        w_id = [obj.work_id for obj in wk_assign]
        status_values = ["no"," "]
        work = Work.objects.filter(id__in=w_id,start_date__gte=st_dt,end_date__lte=end_dt,status__in=status_values).order_by('-start_date')
        print(work)
        for i in work:
            i.difference = i.target - 10

        today = date.today()
        print(today)
        context = {
            "usr":usr,
            # 'dl_task': dl_task,
            "exe":exes,
            # 'warning':warning,
            'eid':exe,
            'work':work,
            'today':today,
            'w_mails':w_mails

        }
    
        return render(request,'team_lead/tl_delay_flt.html', context)
    return redirect("tl_delay_det")

def tl_warning_mail(request,eid):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    c_date=date.today()
    
    if request.method=="POST":
        desc = request.POST.get('description')
        file = request.FILES['file']
        wid = request.POST.get('wid')
        print(wid)
        data = dm_warning_mails(executive_id=eid,work_id=wid,description=desc,date=c_date,file=file)
        data.save()
        work = Work.objects.get(id=wid)
        work.mail_status='yes'
        work.save()

    return redirect("tl_delay_det")

def he_waste_data_home(request):
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    return render(request, 'head/he_waste_data_home.html', {'usr':usr})

def he_waste_data_exe(request):
    ids=request.session['userid'] 
    usr = user_registration.objects.get(id=ids)
    leads_by_date = daily_leeds_exists.objects.values('date').annotate(total_leads=Count('id')).order_by('-date')


    context = {
        "usr":usr,
        'leads_by_date': leads_by_date,
    }
    return render(request, 'head/he_waste_data_exe.html', context)

def he_waste_data_det_exe(request,date=None):
    if request.session.has_key('userid'):
        ids = request.session['userid']
    else:
        return redirect('/')
    ids=request.session['userid']
    usr = user_registration.objects.get(id=ids)
    leads=daily_leeds_exists.objects.filter(date=date).order_by('-date')
   
    context={
        "usr":usr,
        'leads':leads,


    }
    return render(request,'head/he_waste_data_det_exe.html', context)

